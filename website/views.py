
from django.shortcuts import render
from django.http import HttpResponse
from django.http import JsonResponse
from django import template
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.csrf import csrf_protect
from django.http import HttpResponseRedirect
from django.shortcuts import redirect
import json			
import openpyxl
import pymysql
from django.core.exceptions import PermissionDenied
from django.http import HttpResponseBadRequest
from datetime import datetime, date
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
import random
import string
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from validate_email import validate_email
import DNS
import os
from django.conf import settings

from mysite.settings import BASE_DIR


class Content :
    def __init__(self,id='',sid='',ipuid='',ipustart='',ipuend='',ipu='',filename='') :
        self.ipu=ipu
        self.id=id
        self.sid=sid
        self.ipuid=ipuid
        self.ipustart=ipustart
        self.ipuend=ipuend
        self.filename=filename
        self.start_pos = 0
        self.theme =""
        self.sub_content_array=[]
    def get_pos(self,i) :
        self.start_pos = i
    def to_json(self):

        return {
			
            'ipu': self.ipu,
            'filename': self.filename,
	    	'theme': self.theme,
		    "sub_content": self.sub_content_array
        }
class Content_conversation :
    def __init__(self) :
        self.ipu=""
        self.pos = 0
        self.start_time=0
        self.end_time=0
        self.pinyin = 0
        self.word=""
        self.speaker =""
    def to_json(self):
        return {
            'speaker': self.speaker,
            'pinyin': self.pinyin,
            'pos': self.pos,
            'ipu': self.ipu,
        }
class Conversation:
    def __init__(self):
        self.content_array =[]
        self.sub_content_array =[]
        self.theme = ""
        self.talk_id =""

    def add_content(self, content):
        self.content_array.append(content)

    def get_content(self):
         return self.content_array
    def to_json(self):
        conversation_json = []
        for content in self.content_array:
            conversation_json.append(content.to_json()) 
        sub_conversation_json = []
        for each in self.sub_content_array:
            sub_conversation_json.append(each.to_json())
        return {
            "theme": self.theme,
            "talk_id": self.talk_id,
            "content": conversation_json,
            "sub_content": sub_conversation_json
        }
def contain_digital(string1):
  for each in string1:
    if ord(each) >= ord('0') and ord(each) <= ord('9'):
      return True
  return False
def get_chinese_index(num,string1) :
    text = string1[0:num]
    ret = 0
    is_chinese = True
    for each in text :
        if('\u4e00' <= each <= '\u9fff') :
            ret+=1
            is_chinese = True
        else :
            if not (is_chinese) :
               if each ==" ":
                ret +=1
            is_chinese = False
    return ret
def is_chinese_string(input_str):
    for char in input_str:
        if '\u4e00' <= char <= '\u9fff':
            return True
    return False
def get_my_num(inp_arr):
    num=0
    for each in inp_arr :
        if each :
            num+=1
    return num
def pinyinnum2chinese(string1 ,nmu) :
    num2=0
    is_chinese = True
    for each in string1 :
        if nmu == 0 :
            return num2
        num2+=1
        if('\u4e00' <= each <= '\u9fff') :
            nmu-=1
            is_chinese = True
        else :
            if not (is_chinese) :
               if each ==" ":
                nmu -=1
            is_chinese = False
def remove_extra_spaces(input_str):
    result = ' '.join(input_str.split())
    return result
def pinyin2word(word,pinyin) :
    temp = []
    start = 0
    pinyin_split = pinyin.split()
    for each in word :
        if(is_chinese_string(each)) :
            temp_word = ""
            for word in each :
                temp_word+=pinyin_split[start]+" "
                start+=1
            temp_word = temp_word.strip()
            temp.append(temp_word)
        else :
            temp.append(pinyin_split[start])
            start+=1
    return temp
def merge_continuous_english_words(arr):
    merged_arr = []
    current_word = ""

    for word in arr:
        if(is_chinese_string(word)):
            if(current_word!=""):
                 merged_arr.append(current_word)
                 current_word = ""
            merged_arr.append(word)
        else:
            if('(' in word) :
                if(current_word):
                    merged_arr.append(current_word)
                merged_arr.append(word)
                current_word = ""
                continue
            elif(current_word!=""):
                current_word+=" "
            current_word+=word
    if(current_word):
        merged_arr.append(current_word)

    return merged_arr
def get_num(fff) :
    num=0
    is_chinese = False
    for each in fff :
        if('\u4e00' <= each <= '\u9fff') :
            num+=1
            if(is_chinese) :
                num+=1
            is_chinese = False
        else :
            is_chinese = True
    if(is_chinese) :
        num+=1
    return num
def splice(total_word,splice_word,pinyin) :
    my_word = total_word.replace(" ",'')
    my_splice_word = ''.join(splice_word)
    my_splice_word = my_splice_word.replace(" ",'')
	
    pre = my_word.index(my_splice_word)
    post = len(my_word)-pre-len(my_splice_word)
    pre = my_word[0:pre]
    post=my_word[len(my_word)-post:len(my_word)]
    pre_num = get_num(pre)
    post_num = get_num(post)
    ret = pinyin.split(" ")
    ret = ret[pre_num:len(ret)-post_num]
    ret = ' '.join(ret)
    return ret



@csrf_protect
def index(request):
	response = render(request,'index.html',{})
	response['Strict-Transport-Security'] = 'max-age=2592000'
	response['X-Frame-Options'] = 'SAMEORIGIN'
	response['Referrer-Policy'] = 'no-referrer'
	response['X-XSS-Protection'] = '1; mode=block'
	response['X-Content-Type-Options'] = 'nosniff'
	response['Strict-Transport-Security'] = 'max-age=16070400; includeSubDomains'
	return response
@csrf_protect
def corpus_list_ch(request):
	has_session = 'login' in request.session
	context = {
        'has_session': has_session,
    }
	response = render(request,'ch/corpus_list.html',context)
	response['Strict-Transport-Security'] = 'max-age=2592000'
	response['X-Frame-Options'] = 'SAMEORIGIN'
	response['Referrer-Policy'] = 'no-referrer'
	response['X-XSS-Protection'] = '1; mode=block'
	response['X-Content-Type-Options'] = 'nosniff'
	response['Strict-Transport-Security'] = 'max-age=16070400; includeSubDomains'
	return response

@csrf_protect
def corpus_list_en(request):
	has_session = 'login' in request.session
	context = {
        'has_session': has_session,
    }
	response = render(request,'en/corpus_list.html',context)
	response['Strict-Transport-Security'] = 'max-age=2592000'
	response['X-Frame-Options'] = 'SAMEORIGIN'
	response['Referrer-Policy'] = 'no-referrer'
	response['X-XSS-Protection'] = '1; mode=block'
	response['X-Content-Type-Options'] = 'nosniff'
	response['Strict-Transport-Security'] = 'max-age=16070400; includeSubDomains'
	return response

@csrf_protect
def corpus_process_ch(request):
	has_session = 'login' in request.session
	context = {
        'has_session': has_session,
    }
	response = render(request,'ch/corpus_process.html',context)
	response['Strict-Transport-Security'] = 'max-age=2592000'
	response['X-Frame-Options'] = 'SAMEORIGIN'
	response['Referrer-Policy'] = 'no-referrer'
	response['X-XSS-Protection'] = '1; mode=block'
	response['X-Content-Type-Options'] = 'nosniff'
	response['Strict-Transport-Security'] = 'max-age=16070400; includeSubDomains'
	return response

@csrf_protect
def corpus_process_en(request):
	has_session = 'login' in request.session
	context = {
        'has_session': has_session,
    }
	response = render(request,'en/corpus_process.html',context)
	response['Strict-Transport-Security'] = 'max-age=2592000'
	response['X-Frame-Options'] = 'SAMEORIGIN'
	response['Referrer-Policy'] = 'no-referrer'
	response['X-XSS-Protection'] = '1; mode=block'
	response['X-Content-Type-Options'] = 'nosniff'
	response['Strict-Transport-Security'] = 'max-age=16070400; includeSubDomains'
	return response

@csrf_protect
def corpus_resource_ch(request):
	has_session = 'login' in request.session
	context = {
        'has_session': has_session,
    }
	response = render(request,'ch/corpus_resource.html',context)
	response['Strict-Transport-Security'] = 'max-age=2592000'
	response['X-Frame-Options'] = 'SAMEORIGIN'
	response['Referrer-Policy'] = 'no-referrer'
	response['X-XSS-Protection'] = '1; mode=block'
	response['X-Content-Type-Options'] = 'nosniff'
	response['Strict-Transport-Security'] = 'max-age=16070400; includeSubDomains'
	return response

@csrf_protect
def corpus_resource_en(request):
	has_session = 'login' in request.session
	context = {
        'has_session': has_session,
    }
	response = render(request,'en/corpus_resource.html',context)
	response['Strict-Transport-Security'] = 'max-age=2592000'
	response['X-Frame-Options'] = 'SAMEORIGIN'
	response['Referrer-Policy'] = 'no-referrer'
	response['X-XSS-Protection'] = '1; mode=block'
	response['X-Content-Type-Options'] = 'nosniff'
	response['Strict-Transport-Security'] = 'max-age=16070400; includeSubDomains'
	return response

@csrf_protect
def corpus1_ch(request):
	has_session = 'login' in request.session
	context = {
        'has_session': has_session,
    }
	response = render(request,'ch/corpus1.html',context)
	response['Strict-Transport-Security'] = 'max-age=2592000'
	response['X-Frame-Options'] = 'SAMEORIGIN'
	response['Referrer-Policy'] = 'no-referrer'
	response['X-XSS-Protection'] = '1; mode=block'
	response['X-Content-Type-Options'] = 'nosniff'
	response['Strict-Transport-Security'] = 'max-age=16070400; includeSubDomains'
	return response

@csrf_protect
def corpus1_en(request):
	has_session = 'login' in request.session
	context = {
        'has_session': has_session,
    }
	response = render(request,'en/corpus1.html',context)
	response['Strict-Transport-Security'] = 'max-age=2592000'
	response['X-Frame-Options'] = 'SAMEORIGIN'
	response['Referrer-Policy'] = 'no-referrer'
	response['X-XSS-Protection'] = '1; mode=block'
	response['X-Content-Type-Options'] = 'nosniff'
	response['Strict-Transport-Security'] = 'max-age=16070400; includeSubDomains'
	return response

@csrf_protect
def corpus2_ch(request):
	has_session = 'login' in request.session
	context = {
        'has_session': has_session,
    }
	response = render(request,'ch/corpus2.html',context)
	response['Strict-Transport-Security'] = 'max-age=2592000'
	response['X-Frame-Options'] = 'SAMEORIGIN'
	response['Referrer-Policy'] = 'no-referrer'
	response['X-XSS-Protection'] = '1; mode=block'
	response['X-Content-Type-Options'] = 'nosniff'
	response['Strict-Transport-Security'] = 'max-age=16070400; includeSubDomains'
	return response

@csrf_protect
def corpus2_en(request):
	has_session = 'login' in request.session
	context = {
        'has_session': has_session,
    }
	response = render(request,'en/corpus2.html',context)
	response['Strict-Transport-Security'] = 'max-age=2592000'
	response['X-Frame-Options'] = 'SAMEORIGIN'
	response['Referrer-Policy'] = 'no-referrer'
	response['X-XSS-Protection'] = '1; mode=block'
	response['X-Content-Type-Options'] = 'nosniff'
	response['Strict-Transport-Security'] = 'max-age=16070400; includeSubDomains'
	return response

@csrf_protect
def corpus3_ch(request):
	has_session = 'login' in request.session
	context = {
        'has_session': has_session,
    }
	response = render(request,'ch/corpus3.html',context)
	response['Strict-Transport-Security'] = 'max-age=2592000'
	response['X-Frame-Options'] = 'SAMEORIGIN'
	response['Referrer-Policy'] = 'no-referrer'
	response['X-XSS-Protection'] = '1; mode=block'
	response['X-Content-Type-Options'] = 'nosniff'
	response['Strict-Transport-Security'] = 'max-age=16070400; includeSubDomains'
	has_session = 'login' in request.session
	return response

@csrf_protect
def corpus3_en(request):
	has_session = 'login' in request.session
	context = {
        'has_session': has_session,
    }
	response = render(request,'en/corpus3.html',context)
	response['Strict-Transport-Security'] = 'max-age=2592000'
	response['X-Frame-Options'] = 'SAMEORIGIN'
	response['Referrer-Policy'] = 'no-referrer'
	response['X-XSS-Protection'] = '1; mode=block'
	response['X-Content-Type-Options'] = 'nosniff'
	response['Strict-Transport-Security'] = 'max-age=16070400; includeSubDomains'
	return response

@csrf_protect
def corpus4_ch(request):
	has_session = 'login' in request.session
	context = {
        'has_session': has_session,
    }
	response = render(request,'ch/corpus4.html',context)
	response['Strict-Transport-Security'] = 'max-age=2592000'
	response['X-Frame-Options'] = 'SAMEORIGIN'
	response['Referrer-Policy'] = 'no-referrer'
	response['X-XSS-Protection'] = '1; mode=block'
	response['X-Content-Type-Options'] = 'nosniff'
	response['Strict-Transport-Security'] = 'max-age=16070400; includeSubDomains'
	return response

@csrf_protect
def corpus4_en(request):
	has_session = 'login' in request.session
	context = {
        'has_session': has_session,
    }
	response = render(request,'en/corpus4.html',context)
	response['Strict-Transport-Security'] = 'max-age=2592000'
	response['X-Frame-Options'] = 'SAMEORIGIN'
	response['Referrer-Policy'] = 'no-referrer'
	response['X-XSS-Protection'] = '1; mode=block'
	response['X-Content-Type-Options'] = 'nosniff'
	response['Strict-Transport-Security'] = 'max-age=16070400; includeSubDomains'
	return response

@csrf_protect
def corpus5_ch(request):
	has_session = 'login' in request.session
	context = {
        'has_session': has_session,
    }
	response = render(request,'ch/corpus5.html',context)
	response['Strict-Transport-Security'] = 'max-age=2592000'
	response['X-Frame-Options'] = 'SAMEORIGIN'
	response['Referrer-Policy'] = 'no-referrer'
	response['X-XSS-Protection'] = '1; mode=block'
	response['X-Content-Type-Options'] = 'nosniff'
	response['Strict-Transport-Security'] = 'max-age=16070400; includeSubDomains'
	return response

@csrf_protect
def corpus5_en(request):
	has_session = 'login' in request.session
	context = {
        'has_session': has_session,
    }
	response = render(request,'en/corpus5.html',context)
	response['Strict-Transport-Security'] = 'max-age=2592000'
	response['X-Frame-Options'] = 'SAMEORIGIN'
	response['Referrer-Policy'] = 'no-referrer'
	response['X-XSS-Protection'] = '1; mode=block'
	response['X-Content-Type-Options'] = 'nosniff'
	response['Strict-Transport-Security'] = 'max-age=16070400; includeSubDomains'
	return response

@csrf_protect
def resources_ch(request):
	has_session = 'login' in request.session
	context = {
        'has_session': has_session,
    }
	response = render(request,'ch/resources.html',context)
	response['Strict-Transport-Security'] = 'max-age=2592000'
	response['X-Frame-Options'] = 'SAMEORIGIN'
	response['Referrer-Policy'] = 'no-referrer'
	response['X-XSS-Protection'] = '1; mode=block'
	response['X-Content-Type-Options'] = 'nosniff'
	response['Strict-Transport-Security'] = 'max-age=16070400; includeSubDomains'
	return response

@csrf_protect
def resources_en(request):
	has_session = 'login' in request.session
	context = {
        'has_session': has_session,
    }
	response = render(request,'en/resources.html',context)
	response['Strict-Transport-Security'] = 'max-age=2592000'
	response['X-Frame-Options'] = 'SAMEORIGIN'
	response['Referrer-Policy'] = 'no-referrer'
	response['X-XSS-Protection'] = '1; mode=block'
	response['X-Content-Type-Options'] = 'nosniff'
	response['Strict-Transport-Security'] = 'max-age=16070400; includeSubDomains'
	return response

@csrf_protect
def search_ch(request):
	has_session = True
	context = {
        'has_session': has_session,
    }
	response = render(request,'ch/search.html',context)
	response['Strict-Transport-Security'] = 'max-age=2592000'
	response['X-Frame-Options'] = 'SAMEORIGIN'
	response['Referrer-Policy'] = 'no-referrer'
	response['X-XSS-Protection'] = '1; mode=block'
	response['X-Content-Type-Options'] = 'nosniff'
	response['Strict-Transport-Security'] = 'max-age=16070400; includeSubDomains'
	return response

@csrf_protect
def search_en(request):
	has_session = 'login' in request.session
	context = {
        'has_session': has_session,
    }
	response = render(request,'en/search.html',context)
	response['Strict-Transport-Security'] = 'max-age=2592000'
	response['X-Frame-Options'] = 'SAMEORIGIN'
	response['Referrer-Policy'] = 'no-referrer'
	response['X-XSS-Protection'] = '1; mode=block'
	response['X-Content-Type-Options'] = 'nosniff'
	response['Strict-Transport-Security'] = 'max-age=16070400; includeSubDomains'
	return response

@csrf_protect
def search2_ch(request):
	has_session = 'login' in request.session
	context = {
        'has_session': has_session,
    }
	response = render(request,'ch/search2.html',context)
	response['Strict-Transport-Security'] = 'max-age=2592000'
	response['X-Frame-Options'] = 'SAMEORIGIN'
	response['Referrer-Policy'] = 'no-referrer'
	response['X-XSS-Protection'] = '1; mode=block'
	response['X-Content-Type-Options'] = 'nosniff'
	response['Strict-Transport-Security'] = 'max-age=16070400; includeSubDomains'
	return response

@csrf_protect
def search2_en(request):
	has_session = 'login' in request.session
	context = {
        'has_session': has_session,
    }
	response = render(request,'en/search2.html',context)
	response['Strict-Transport-Security'] = 'max-age=2592000'
	response['X-Frame-Options'] = 'SAMEORIGIN'
	response['Referrer-Policy'] = 'no-referrer'
	response['X-XSS-Protection'] = '1; mode=block'
	response['X-Content-Type-Options'] = 'nosniff'
	response['Strict-Transport-Security'] = 'max-age=16070400; includeSubDomains'
	return response

@csrf_protect
def signin_ch(request):
	has_session = 'login' in request.session
	context = {
        'has_session': has_session,
    }
	response = render(request,'ch/signin.html',context)
	response['Strict-Transport-Security'] = 'max-age=2592000'
	response['X-Frame-Options'] = 'SAMEORIGIN'
	response['Referrer-Policy'] = 'no-referrer'
	response['X-XSS-Protection'] = '1; mode=block'
	response['X-Content-Type-Options'] = 'nosniff'
	response['Strict-Transport-Security'] = 'max-age=16070400; includeSubDomains'
	return response

@csrf_protect
def signin_en(request):
	has_session = 'login' in request.session
	context = {
        'has_session': has_session,
    }
	response = render(request,'en/signin.html',context)
	response['Strict-Transport-Security'] = 'max-age=2592000'
	response['X-Frame-Options'] = 'SAMEORIGIN'
	response['Referrer-Policy'] = 'no-referrer'
	response['X-XSS-Protection'] = '1; mode=block'
	response['X-Content-Type-Options'] = 'nosniff'
	response['Strict-Transport-Security'] = 'max-age=16070400; includeSubDomains'
	return response

@csrf_protect
def account_ch(request):
	has_session = 'login' in request.session
	context = {
        'has_session': has_session,
    }
	response = render(request,'ch/account.html',context)
	response['Strict-Transport-Security'] = 'max-age=2592000'
	response['X-Frame-Options'] = 'SAMEORIGIN'
	response['Referrer-Policy'] = 'no-referrer'
	response['X-XSS-Protection'] = '1; mode=block'
	response['X-Content-Type-Options'] = 'nosniff'
	response['Strict-Transport-Security'] = 'max-age=16070400; includeSubDomains'
	return response

@csrf_protect
def account_en(request):
	has_session = 'login' in request.session
	context = {
        'has_session': has_session,
    }
	response = render(request,'en/account.html',context)
	response['Strict-Transport-Security'] = 'max-age=2592000'
	response['X-Frame-Options'] = 'SAMEORIGIN'
	response['Referrer-Policy'] = 'no-referrer'
	response['X-XSS-Protection'] = '1; mode=block'
	response['X-Content-Type-Options'] = 'nosniff'
	response['Strict-Transport-Security'] = 'max-age=16070400; includeSubDomains'
	return response

@csrf_protect
def apply_ch(request):
	has_session = 'login' in request.session
	context = {
        'has_session': has_session,
    }
	response = render(request,'ch/apply.html',context)
	response['Strict-Transport-Security'] = 'max-age=2592000'
	response['X-Frame-Options'] = 'SAMEORIGIN'
	response['Referrer-Policy'] = 'no-referrer'
	response['X-XSS-Protection'] = '1; mode=block'
	response['X-Content-Type-Options'] = 'nosniff'
	response['Strict-Transport-Security'] = 'max-age=16070400; includeSubDomains'
	return response

@csrf_protect
def apply_en(request):
	has_session = 'login' in request.session
	context = {
        'has_session': has_session,
    }
	response = render(request,'en/apply.html',context)
	response['Strict-Transport-Security'] = 'max-age=2592000'
	response['X-Frame-Options'] = 'SAMEORIGIN'
	response['Referrer-Policy'] = 'no-referrer'
	response['X-XSS-Protection'] = '1; mode=block'
	response['X-Content-Type-Options'] = 'nosniff'
	response['Strict-Transport-Security'] = 'max-age=16070400; includeSubDomains'
	return response

@csrf_protect
def copyright_ch(request):
	has_session = 'login' in request.session
	context = {
        'has_session': has_session,
    }
	response = render(request,'ch/copyright.html',context)
	response['Strict-Transport-Security'] = 'max-age=2592000'
	response['X-Frame-Options'] = 'SAMEORIGIN'
	response['Referrer-Policy'] = 'no-referrer'
	response['X-XSS-Protection'] = '1; mode=block'
	response['X-Content-Type-Options'] = 'nosniff'
	response['Strict-Transport-Security'] = 'max-age=16070400; includeSubDomains'
	return response

@csrf_protect
def copyright_en(request):
	has_session = 'login' in request.session
	context = {
        'has_session': has_session,
    }
	response = render(request,'en/copyright.html',context)
	response['Strict-Transport-Security'] = 'max-age=2592000'
	response['X-Frame-Options'] = 'SAMEORIGIN'
	response['Referrer-Policy'] = 'no-referrer'
	response['X-XSS-Protection'] = '1; mode=block'
	response['X-Content-Type-Options'] = 'nosniff'
	response['Strict-Transport-Security'] = 'max-age=16070400; includeSubDomains'
	return response

@csrf_protect
def guide_ch(request):
	has_session = 'login' in request.session
	context = {
        'has_session': has_session,
    }
	response = render(request,'ch/guide.html',context)
	response['Strict-Transport-Security'] = 'max-age=2592000'
	response['X-Frame-Options'] = 'SAMEORIGIN'
	response['Referrer-Policy'] = 'no-referrer'
	response['X-XSS-Protection'] = '1; mode=block'
	response['X-Content-Type-Options'] = 'nosniff'
	response['Strict-Transport-Security'] = 'max-age=16070400; includeSubDomains'
	return response

@csrf_protect
def guide_en(request):
	has_session = 'login' in request.session
	context = {
        'has_session': has_session,
    }
	response = render(request,'en/guide.html',context)
	response['Strict-Transport-Security'] = 'max-age=2592000'
	response['X-Frame-Options'] = 'SAMEORIGIN'
	response['Referrer-Policy'] = 'no-referrer'
	response['X-XSS-Protection'] = '1; mode=block'
	response['X-Content-Type-Options'] = 'nosniff'
	response['Strict-Transport-Security'] = 'max-age=16070400; includeSubDomains'
	return response

@csrf_protect
def forgotpw_ch(request):
	has_session = 'login' in request.session
	context = {
        'has_session': has_session,
    }
	response = render(request,'ch/forgotpw.html',context)
	response['Strict-Transport-Security'] = 'max-age=2592000'
	response['X-Frame-Options'] = 'SAMEORIGIN'
	response['Referrer-Policy'] = 'no-referrer'
	response['X-XSS-Protection'] = '1; mode=block'
	response['X-Content-Type-Options'] = 'nosniff'
	response['Strict-Transport-Security'] = 'max-age=16070400; includeSubDomains'
	return response

@csrf_protect
def forgotpw_en(request):
	has_session = 'login' in request.session
	context = {
        'has_session': has_session,
    }
	response = render(request,'en/forgotpw.html',context)
	response['Strict-Transport-Security'] = 'max-age=2592000'
	response['X-Frame-Options'] = 'SAMEORIGIN'
	response['Referrer-Policy'] = 'no-referrer'
	response['X-XSS-Protection'] = '1; mode=block'
	response['X-Content-Type-Options'] = 'nosniff'
	response['Strict-Transport-Security'] = 'max-age=16070400; includeSubDomains'
	return response

@csrf_protect
def backend_index(request):
	response = render(request,'backend/index.html',{})
	response['Strict-Transport-Security'] = 'max-age=2592000'
	response['X-Frame-Options'] = 'SAMEORIGIN'
	response['Referrer-Policy'] = 'no-referrer'
	response['X-XSS-Protection'] = '1; mode=block'
	response['X-Content-Type-Options'] = 'nosniff'
	response['Strict-Transport-Security'] = 'max-age=16070400; includeSubDomains'
	return response

@csrf_protect
def backend_record(request):
	if 'backend_login' not in request.session:
		return HttpResponseBadRequest('權限不足，請先登錄')
	response = render(request,'backend/record.html',{})
	response['Strict-Transport-Security'] = 'max-age=2592000'
	response['X-Frame-Options'] = 'SAMEORIGIN'
	response['Referrer-Policy'] = 'no-referrer'
	response['X-XSS-Protection'] = '1; mode=block'
	response['X-Content-Type-Options'] = 'nosniff'
	response['Strict-Transport-Security'] = 'max-age=16070400; includeSubDomains'
	return response

@csrf_protect
def backend_corpus(request):
	if 'backend_login' not in request.session:
		return HttpResponseBadRequest('權限不足，請先登錄')
	response = render(request,'backend/corpus.html',{})
	response['Strict-Transport-Security'] = 'max-age=2592000'
	response['X-Frame-Options'] = 'SAMEORIGIN'
	response['Referrer-Policy'] = 'no-referrer'
	response['X-XSS-Protection'] = '1; mode=block'
	response['X-Content-Type-Options'] = 'nosniff'
	response['Strict-Transport-Security'] = 'max-age=16070400; includeSubDomains'
	return response

@csrf_protect
def backend_account(request):
	if 'backend_login' not in request.session:
		return HttpResponseBadRequest('權限不足，請先登錄')
	response = render(request,'backend/account.html',{})
	response['Strict-Transport-Security'] = 'max-age=2592000'
	response['X-Frame-Options'] = 'SAMEORIGIN'
	response['Referrer-Policy'] = 'no-referrer'
	response['X-XSS-Protection'] = '1; mode=block'
	response['X-Content-Type-Options'] = 'nosniff'
	response['Strict-Transport-Security'] = 'max-age=16070400; includeSubDomains'
	return response

@csrf_protect
def backend_manual(request):
	if 'backend_login' not in request.session:
		return HttpResponseBadRequest('權限不足，請先登錄')
	response = render(request,'backend/manual.html',{})
	response['Strict-Transport-Security'] = 'max-age=2592000'
	response['X-Frame-Options'] = 'SAMEORIGIN'
	response['Referrer-Policy'] = 'no-referrer'
	response['X-XSS-Protection'] = '1; mode=block'
	response['X-Content-Type-Options'] = 'nosniff'
	response['Strict-Transport-Security'] = 'max-age=16070400; includeSubDomains'
	return response

@csrf_protect
def backend_license(request):
	if 'backend_login' not in request.session:
		return HttpResponseBadRequest('權限不足，請先登錄')
	response = render(request,'backend/license.html',{})
	response['Strict-Transport-Security'] = 'max-age=2592000'
	response['X-Frame-Options'] = 'SAMEORIGIN'
	response['Referrer-Policy'] = 'no-referrer'
	response['X-XSS-Protection'] = '1; mode=block'
	response['X-Content-Type-Options'] = 'nosniff'
	response['Strict-Transport-Security'] = 'max-age=16070400; includeSubDomains'
	return response

@csrf_protect
def backend_resource(request):
	if 'backend_login' not in request.session:
		return HttpResponseBadRequest('權限不足，請先登錄')
	response = render(request,'backend/resource.html',{})
	response['Strict-Transport-Security'] = 'max-age=2592000'
	response['X-Frame-Options'] = 'SAMEORIGIN'
	response['Referrer-Policy'] = 'no-referrer'
	response['X-XSS-Protection'] = '1; mode=block'
	response['X-Content-Type-Options'] = 'nosniff'
	response['Strict-Transport-Security'] = 'max-age=16070400; includeSubDomains'
	return response

@csrf_exempt
def login(request):
	result = {'status':'','msg':'','data':{}}
	with open('static/media/link.json', 'r', encoding='utf-16') as file:
		json_data = json.load(file)
	db = pymysql.connect(host=json_data['db']['host'],
                     user=json_data['db']['user'],
                     password=json_data['db']['passwd'],
                     database=json_data['db']['db'])
	cursor = db.cursor()
	try:
		account = request.POST['account']
		pwd = request.POST['pwd']

		if check_illegal_parameter(account) or check_illegal_parameter(pwd):
			result['status'] = 'No'
			result['msg'] = 'illegal_parameter'
		else:
			sql = "SELECT * FROM `Member` WHERE `account`='%s' AND pwd='%s';"%(account,pwd)
			cursor.execute(sql)
			has_account = cursor.fetchall()
			if has_account :
				result['status'] = 'Yes'
				request.session['login'] = account
			else :
				result['status'] = 'No'
				result['msg'] = 'error_account_or_password'
	except Exception as e:
		result['status'] = 'Error'
		result['msg'] = str(e)
	return JsonResponse(result, safe=False)

@csrf_exempt
def backend_login(request):
	result = {'status':'','msg':'','data':{}}
	with open('static/media/link.json', 'r', encoding='utf-16') as file:
		json_data = json.load(file)
	db = pymysql.connect(host=json_data['db']['host'],
                     user=json_data['db']['user'],
                     password=json_data['db']['passwd'],
                     database=json_data['db']['db'])
	cursor = db.cursor()

	try:
		account = request.POST['account']
		pwd = request.POST['pwd']
	
		if check_illegal_parameter(account) or check_illegal_parameter(pwd):
			result['status'] = 'No'
			result['msg'] = 'illegal_parameter'
		
		else:
			sql = "SELECT * FROM `Administrator` WHERE `account`='%s' AND pwd='%s';"%(account,pwd)
			cursor.execute(sql)
			has_account = cursor.fetchall()
			if has_account :
				result['status'] = 'Yes'
				request.session['backend_login'] = account
			else :
				result['status'] = 'No'
				result['msg'] = 'error_account_or_password'
	except Exception as e:
		result['status'] = 'Error'
		result['msg'] = str(e)
	db.close()
	return JsonResponse(result, safe=False)

@csrf_exempt
def member_logout(request):
	if 'login' in request.session:
		del request.session['login']
		request.session.modified = True
		return JsonResponse({'message': '成功'})
	else:
		return JsonResponse({'message': '失敗'})

@csrf_exempt
def backend_logout(request):
	result = {}
	try:
		# Check if the user is logged in
		if 'backend_login' in request.session:
		# User is logged in, perform logout
			del request.session['backend_login']
		result['status'] = 'Yes'
	except Exception as e:
		result['status'] = 'Error'
		result['msg'] = str(e)
	return JsonResponse(result, safe=False)
@csrf_exempt
def backend_download_corpus(request):
	if 'backend_login' not in request.session:
		raise PermissionDenied('無效的請求方法')

	# 連接資料庫
	with open('static/media/link.json', 'r', encoding='utf-16') as file:
		json_data = json.load(file)
	db = pymysql.connect(host=json_data['db']['host'],
                     user=json_data['db']['user'],
                     password=json_data['db']['passwd'],
                     database=json_data['db']['db'])
	cursor = db.cursor()
	sql = """
	SELECT 
  (CASE WHEN COUNT(t1.id) > 0 AND COUNT(t2.id) > 0 AND COUNT(t3.id) > 0 THEN 'All tables have values' ELSE 'At least one table is empty' END) AS result
	FROM
	Socio t1
	LEFT JOIN Modern_Chinese t2 ON 1=1
	LEFT JOIN Child t3 ON 1=1;
	"""
	cursor.execute(sql)
	result = cursor.fetchone()

	if(result[0] =='At least one table is empty') :
		db.close()
		return JsonResponse({'message': '資料表遺失，請重新上傳'})
	# 創建一個新的工作簿
	workbook = openpyxl.Workbook()


	# 刪除預設的工作表
	default_sheet = workbook['Sheet']
	workbook.remove(default_sheet)
	#寫入現代漢語對話進xlxs
	MC = ['MCDC_8','MCDC_22','MTCC','MMTC']
	for each in MC :
		# 創建一個工作表
		worksheet = workbook.create_sheet(title=each)
		# 寫入標題行設定間距
		column_widths = {'A': 13.83, 'B': 20.67, 'C': 20.83, 'D': 18, 'E': 11.5, 'F': 13.83, 'G': 17.67, 'H': 18.67, 'I': 58.67, 'J': 11.17, 'K': 7.83}  # 列寬度字典，键為列字母，值為宽度
		headers = ['對話編號', '語者編號', 'Filename', 'FileSN', 'IPUID', 'IPUStart', 'IPUEnd', 'IPU', 'Word', 'POS', 'Pinyin']
		for col, header in enumerate(headers, start=1):
			column_letter = openpyxl.utils.get_column_letter(col)
			worksheet.cell(row=1, column=col, value=header)
			if column_letter in column_widths:
				worksheet.column_dimensions[column_letter].width = column_widths[column_letter]
		#索引對應主題的資料
		sql = "SELECT * FROM `Modern_Chinese` WHERE `theme` = '%s'"%each
		cursor.execute(sql)
		results = cursor.fetchall()
		sql2 = ''
		# 用字典記錄每一筆的talkid speakerid 並記錄包含的檔案id
		mydict = dict()
	
		for row in results: 
			if sql2 == '' :
				sql2 += "WHERE MH_id=%s"%row[0]
			else :
				sql2 += " OR MH_id=%s"%row[0]
			mydict["%stalkid"%row[0]] = row[2]
			mydict["%sspeakerid"%row[0]] = row[3]
		
		#索引檔案
		sql = "SELECT * FROM `MH_File`" +sql2 +"  ORDER BY FileSn ASC, IPUStart ASC;"
		cursor.execute(sql)
		results = cursor.fetchall()
		#檔案寫入到資料表中
		row_data = []
		for row in results:
			row_data.append([mydict["%stalkid"%row[1]],  mydict["%sspeakerid"%row[1]], row[2], row[3], row[4], row[5], row[6], row[7], row[8], row[9]])


		for row in row_data:
			worksheet.append(row)

		
	#寫入社會語音資料庫進xlxs
	
	sql = "SELECT * FROM `Socio` ;"
	cursor.execute(sql)
	results = cursor.fetchall()
	sql2 = ''
	# 創建一個工作表
	worksheet = workbook.create_sheet(title='Socio')
	# 寫入標題行設定間距
	column_widths = {'A': 10.67, 'B': 11.67, 'C': 7.83, 'D': 20.17, 'E': 7.83, 'F': 43.67}  # 列寬度字典，键為列字母，值為宽度
	headers = ['語者編號', 'Filename',  'IPUID', 'IPUStart', 'IPUEnd', 'IPU']
	for col, header in enumerate(headers, start=1):
		column_letter = openpyxl.utils.get_column_letter(col)
		worksheet.cell(row=1, column=col, value=header)
		if column_letter in column_widths:
			worksheet.column_dimensions[column_letter].width = column_widths[column_letter]
	# 用字典記錄每一筆的 speakerid 
	mydict = dict()
	for row in results: 
		mydict["%sspeakerid"%row[0]] = row[1]
	# 索引檔案
	sql = "SELECT * FROM `S_File` ORDER BY `Filename` ASC , IPUStart ASC;"
	cursor.execute(sql)
	results = cursor.fetchall()
	row_data = []
	#寫入到資料表中
	for row in results:
		row_data.append([ mydict["%sspeakerid"%row[1]], row[6], 
				row[2], row[3], row[4], row[5]])
	for row in row_data:
		worksheet.append(row)



	#寫入兒童對話進xlxs
	child=['Child_HI','Child_NH']
	for each in child :
		# 創建一個工作表
		worksheet = workbook.create_sheet(title=each)
		# 寫入標題行設定間距
		column_widths = {'A': 11.67, 'B': 7.17, 'C': 11.83, 'D': 11.83, 'E': 59.50 }  # 列寬度字典，键為列字母，值為宽度
		headers = ['語者編號',  'IPUID', 'IPUStart', 'IPUEnd', 'IPU']
		for col, header in enumerate(headers, start=1):
			column_letter = openpyxl.utils.get_column_letter(col)
			worksheet.cell(row=1, column=col, value=header)
			if column_letter in column_widths:
				worksheet.column_dimensions[column_letter].width = column_widths[column_letter]
		# 索引對應主題的資料 
		mydict = dict()
		sql = "SELECT * FROM `Child` WHERE `theme` = '%s' "%each
		cursor.execute(sql)
		results = cursor.fetchall()
		sql2 = ''
		# 用字典記錄每一筆的 speakerid 並記錄下id
		for row in results: 
			if sql2 == '' :
				sql2 += "WHERE Cid=%s"%row[0]
			else :
				sql2 += " OR Cid=%s"%row[0]
			mydict["%sspeakerid"%row[0]] = row[2]
		# 索引檔案
		sql = "SELECT * FROM `Child_File`"+sql2 +" ORDER BY `Cid` ASC , IPUStart ASC;"
		cursor.execute(sql)
		results = cursor.fetchall()
		row_data = []
		#寫入到資料表中
		for row in results:
			row_data.append([ mydict["%sspeakerid"%row[1]], 
					row[2], row[3], row[4], row[5]])
		for row in row_data:
			worksheet.append(row)
    # 設定下載檔案的檔名	
    
	file_name = 'Sinca_Corpora.xlsx'
    # 將Excel工作簿轉換為二進位數據
    
	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
	response['Content-Disposition'] = 'attachment; filename="{}"'.format(file_name)
    
	record("下載語料檔",request.session['backend_login'],1)
	workbook.save(response)

    
	return response
@csrf_exempt
def backend_update_corpus(request):
	if 'backend_login' not in request.session:
		raise PermissionDenied('無效的請求方法')

	current_datetime = datetime.now()
	date_string = current_datetime.strftime("%Y-%m-%d %H:%M:%S")
	# 創建資料庫連接
	with open('static/media/link.json', 'r', encoding='utf-16') as file:
		json_data = json.load(file)
	db = pymysql.connect(host=json_data['db']['host'],
                     user=json_data['db']['user'],
                     password=json_data['db']['passwd'],
                     database=json_data['db']['db'])
	cursor = db.cursor()
	sql = """
	SELECT 
  	(CASE WHEN COUNT(t1.id) > 0 OR COUNT(t2.id) > 0 OR COUNT(t3.id) > 0 THEN 'Tables have values' ELSE 'Table is empty' END) AS result
	FROM
	Socio t1
	LEFT JOIN Modern_Chinese t2 ON 1=1
	LEFT JOIN Child t3 ON 1=1;
	"""
	cursor.execute(sql)
	result = cursor.fetchone()

	if(result[0] =='Tables have values') :
		db.close()
		return JsonResponse({'message': '資料表已有資料，請先刪除在上傳'})
	corpus_file = request.FILES.get('corpus_file')

	# 打開文件
	xlsx = pd.ExcelFile(corpus_file)
	try :
		current_datetime = datetime.now()
		date_string = current_datetime.strftime("%Y-%m-%d %H:%M:%S")

		sheet_list=['MCDC_8','MCDC_22','MTCC','MMTC']
		id = 1
		data = []
		data2=[]
		for each in sheet_list :
			df1 = pd.read_excel(xlsx, sheet_name=each )
			df1.fillna(value='NA', inplace=True)   
			unique_rows = df1.drop_duplicates(subset='語者編號')
			for index, row in unique_rows.iterrows():
				current_datetime = datetime.now()
				date_string = current_datetime.strftime("%Y-%m-%d %H:%M:%S")
				conversation_id = row['對話編號']
				speaker_id = row['語者編號']
				sql = "INSERT INTO Modern_Chinese (id,theme, Talk_id, Speaker_id,Log_time) VALUES (%s,%s, %s, %s,%s)"
				data2.append((id,each,conversation_id,speaker_id,date_string))
				filtered_rows = df1[df1['語者編號'] == speaker_id]
				for index, row in filtered_rows.iterrows():
					filename = row['Filename']
					filesn = row['FileSN']
					ipuid = row['IPUID']
					IPUStart = row['IPUStart']
					IPUEnd = row['IPUEnd']
					IPU = row['IPU']
					Word = row['Word']
					POS = row['POS']
					Pinyin = row['Pinyin']
					data.append((id,filename,filesn,ipuid,IPUStart,IPUEnd,IPU,Word,POS,Pinyin,date_string))
				id+=1
		sql = "INSERT INTO Modern_Chinese (id,theme, Talk_id, Speaker_id,Log_time) VALUES (%s,%s, %s, %s,%s)"
		cursor.executemany(sql, data2)
		sql = "INSERT INTO MH_File(MH_id, Filename, FileSn, IPUID, IPUStart, IPUEnd, IPU, Word, POS, Pinyin, Log_time) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"  
		cursor.executemany(sql, data)
		id=1
		data=[]
		data2=[] 
		df1 = pd.read_excel(xlsx, sheet_name='Socio' )
		df1.fillna(value='NA', inplace=True)      
		unique_rows = df1.drop_duplicates(subset='語者編號')
		for index, row in unique_rows.iterrows():
			speaker_id = row['語者編號']
			data2.append((id,speaker_id,date_string))
			filtered_rows = df1[df1['語者編號'] == speaker_id]
			for index, row in filtered_rows.iterrows():
				ipuid = row['IPUID']
				IPUStart = row['IPUStart']
				IPUEnd = row['IPUEnd']
				IPU = row['IPU']
				filename = row['Filename']
				data.append((id,ipuid,IPUStart,IPUEnd,IPU,filename,date_string))    
			id+=1
		sql = "INSERT INTO Socio (id, Speaker_id,Log_time) VALUES (%s,  %s,%s)"
		cursor.executemany(sql, data2)
		sql = "INSERT INTO S_File(Sid,IPUID, IPUStart, IPUEnd, IPU ,Filename,Log_time) VALUES (%s, %s, %s, %s, %s,%s, %s)"
		cursor.executemany(sql, data)
		id=1
		data=[]
		data2=[]
		sheet_list=['Child_HI','Child_NH'] 
		for  each in sheet_list :
			df1 = pd.read_excel(xlsx, sheet_name=each )
			df1.fillna(value='NA', inplace=True)
			unique_rows = df1.drop_duplicates(subset='語者編號')
			for index, row in unique_rows.iterrows():
				speaker_id = row['語者編號']
				data2.append((id,each,speaker_id,date_string))
				filtered_rows = df1[df1['語者編號'] == speaker_id]
				for index, row in filtered_rows.iterrows():
					ipuid = row['IPUID']
					IPUStart = row['IPUStart']
					IPUEnd = row['IPUEnd']
					IPU = row['IPU']
					data.append((id,ipuid,IPUStart,IPUEnd,IPU,date_string))
				id+=1
		sql = "INSERT INTO Child (id,theme, Speaker_id,Log_time) VALUES ( %s,%s, %s,%s)"
		cursor.executemany(sql, data2)
		sql = "INSERT INTO Child_File(Cid,IPUID, IPUStart, IPUEnd, IPU ,Log_time) VALUES (%s, %s, %s, %s, %s, %s)"       
		cursor.executemany(sql, data)
		db.commit()        
		cursor.close()
		db.close()
		record("上傳語料檔",request.session['backend_login'],1)
		return JsonResponse({'message': '新增成功'})
	except Exception as e:

		return JsonResponse({'message': '上傳檔案錯誤，請確認檔案格式 '+str(e) })
@csrf_exempt
def backend_delete_corpus(request) :
	if 'backend_login' not in request.session:
		raise PermissionDenied('無效的請求方法')

	with open('static/media/link.json', 'r', encoding='utf-16') as file:
		json_data = json.load(file)
	db = pymysql.connect(host=json_data['db']['host'],
                     user=json_data['db']['user'],
                     password=json_data['db']['passwd'],
                     database=json_data['db']['db'])

	cursor = db.cursor()
	sql = """
	SELECT 
  	(CASE WHEN COUNT(t1.id) > 0 OR COUNT(t2.id) > 0 OR COUNT(t3.id) > 0 THEN ' Tables have values' ELSE 'Table is empty' END) AS result
	FROM
	Socio t1
	LEFT JOIN Modern_Chinese t2 ON 1=1
	LEFT JOIN Child t3 ON 1=1;
	"""
	cursor.execute(sql)
	result = cursor.fetchone()

	if(result[0] =='Table is empty') :
		db.close()
		return JsonResponse({'message': '資料表已為空'})
	cursor.execute("DELETE FROM `MH_File`")
	cursor.execute("DELETE FROM `Modern_Chinese`")
	cursor.execute("DELETE FROM `S_File`")
	cursor.execute("DELETE FROM `Socio`")
	cursor.execute("DELETE FROM `Child_File`")
	cursor.execute("DELETE FROM `Child`")
	db.commit()
	cursor.close()
	db.close()
	record("刪除語料檔",request.session['backend_login'],1)
	return JsonResponse({'message': '刪除成功'})
@csrf_exempt
def	backend_Get_user_log(request) :
	if 'backend_login' not in request.session:
		raise PermissionDenied('無效的請求方法')
	with open('static/media/link.json', 'r', encoding='utf-16') as file:
		json_data = json.load(file)
	db = pymysql.connect(host=json_data['db']['host'],
					user=json_data['db']['user'],
					password=json_data['db']['passwd'],
					database=json_data['db']['db'])
	cursor = db.cursor()
	sql = """
	SELECT
		User_Log.Is_Administrator,
		CASE
			WHEN Is_Administrator = 1 THEN Administrator.institution
			ELSE member.institution
		END AS institution,
		CASE
			WHEN Is_Administrator = 1 THEN Administrator.account
			ELSE member.account
		END AS account,
		CASE
			WHEN Is_Administrator = 1 THEN Administrator.name
			ELSE member.name
		END AS name,
		User_Log.time,
		User_Log.operate
	FROM
		User_Log
	LEFT JOIN
		Administrator ON User_Log.User_id = Administrator.id
	LEFT JOIN
		member ON User_Log.User_id = member.id;
	"""
	cursor.execute(sql)
	results = cursor.fetchall()
	data = []
	for row in results:
		if row[0] :
			Position = '管理員'
		else :
			Position = '會員'
		data.append({
			'position': Position,
			'institution': row[1],
			'account':  row[2],
			'name': row[3],
			'time': row[4],
			'operate': row[5],
		})
	return JsonResponse(data, safe=False)
@csrf_exempt
def backend_check_email(request) :
	if 'backend_login' not in request.session:
		raise PermissionDenied('無效的請求方法')
	result = {'status':'','msg':''}
	with open('static/media/link.json', 'r', encoding='utf-16') as file:
		json_data = json.load(file)
	db = pymysql.connect(host=json_data['db']['host'],
                     user=json_data['db']['user'],
                     password=json_data['db']['passwd'],
                     database=json_data['db']['db'])
	cursor = db.cursor()
	email = request.POST['email']
	try:
		if check_illegal_parameter(email):
			result['status'] = 'No'
			result['msg'] = '不合法字元'
			return JsonResponse(result, safe=False)
		else:
			sql = "SELECT * FROM `Member` WHERE `email`='%s';"%email
			cursor.execute(sql)
			has_email = cursor.fetchall()
			if has_email :
				result['status'] = 'No'
				result['msg'] = "信箱重複"
				return JsonResponse(result, safe=False)
	except Exception as e:
		result['status'] = 'Error'
		result['msg'] = str(e)
	db.close()
	if result['msg']=='No' :
		return JsonResponse(result, safe=False)
	valid  = validate_email(email,check_mx=True ,verify=True,smtp_timeout=5)
	if (valid):
		result['status'] = 'Yes'
		result['msg'] = "信箱可用"
	else :
		result['status'] = 'No'
		result['msg'] = "信箱無效"
		

	return JsonResponse(result, safe=False)
@csrf_exempt
def backend_check_account(request) :
	if 'backend_login' not in request.session:
		raise PermissionDenied('無效的請求方法')
	
	result = {'status':'','msg':''}
	with open('static/media/link.json', 'r', encoding='utf-16') as file:
		json_data = json.load(file)
	db = pymysql.connect(host=json_data['db']['host'],
                     user=json_data['db']['user'],
                     password=json_data['db']['passwd'],
                     database=json_data['db']['db'])
	cursor = db.cursor()
	try:
		account = request.POST['account']
		if check_illegal_parameter(account):
			result['status'] = 'No'
			result['msg'] = '不合法字元'
		else:
			sql = "SELECT * FROM `Member` WHERE `account`='%s';"%account
			cursor.execute(sql)
			has_email = cursor.fetchall()
			if has_email :
				result['status'] = 'No'
				result['msg'] = "帳號重複"
			else :
				result['status'] = 'Yes'
				result['msg'] = '帳號可用'
	except Exception as e:
		result['status'] = 'Error'
		result['msg'] = str(e)
	db.close()
	return JsonResponse(result, safe=False)
@csrf_exempt
def backend_create_account(request):
	if 'backend_login' not in request.session:
		raise PermissionDenied('無效的請求方法')
	result = {'status':'','msg':''}
	with open('static/media/link.json', 'r', encoding='utf-16') as file:
		json_data = json.load(file)
	db = pymysql.connect(host=json_data['db']['host'],
                     user=json_data['db']['user'],
                     password=json_data['db']['passwd'],
                     database=json_data['db']['db'])
	cursor = db.cursor()

	try:
		email = request.POST['email']
		name = request.POST['name']
		account = request.POST['account']
		phone = request.POST['phone']
		institution = request.POST['institution']
		EffectiveDate = request.POST['EffectiveDate']
		ExpiryDate = request.POST['ExpiryDate']
		current_datetime = datetime.now()
		date_string = current_datetime.strftime("%Y-%m-%d %H:%M:%S")
		if check_illegal_parameter(account) or check_illegal_parameter(email) or check_illegal_parameter(name) or check_illegal_parameter(phone) or check_illegal_parameter(institution)  :
			result['status'] = 'No'
			result['msg'] = 'illegal_parameter'

		
		else:
			characters = string.ascii_letters + string.digits

			# 生成 8 碼的隨機亂碼
			password = ''.join(random.choice(characters) for _ in range(8))
			subject = "語料庫密碼"
			body = "您的密碼是 : %s"%password
			recipients = ["%s"%email]
			send_email(subject, body, recipients)
			sql = "INSERT INTO `Member` (`account`, `pwd`, `name`, `email`, `phone`, `institution`, `Log_time`) VALUES ('%s', '%s', '%s', '%s', '%s', '%s', '%s')"%(account, password, name, email, phone, institution, date_string)
			cursor.execute(sql)
			db.commit()
			last_insert_id = cursor.lastrowid
			sql = "INSERT INTO `Member_Log`(`Member_id`, `start_time`, `end_time`, `Log_time`) VALUES ('%s','%s','%s','%s')"%(last_insert_id,EffectiveDate,ExpiryDate,date_string)
			cursor.execute(sql)
			db.commit()
			result['status'] = 'Yes'

	except Exception as e:
		result['status'] = 'Error'
		result['msg'] = str(e)
	db.close()
	record("新增會員",request.session['backend_login'],1)
	return JsonResponse(result, safe=False)
def check_illegal_parameter(parameter):
	sign_list = ["=","%","+","$","*","/","#","!","?","^","&","<",">","'",'"']
	for sign in sign_list:
		if sign in parameter:
			return True
	return False
def record(way,account,is_administrator) :
	with open('static/media/link.json', 'r', encoding='utf-16') as file:
		json_data = json.load(file)
	db = pymysql.connect(host=json_data['db']['host'],
                     user=json_data['db']['user'],
                     password=json_data['db']['passwd'],
                     database=json_data['db']['db'])
	cursor = db.cursor()
	if is_administrator:
		table_name = 'Administrator'
	else:
		table_name = 'Member'
	sql = "SELECT `id` FROM `%s` WHERE account = '%s'"%(table_name,account)
	cursor.execute(sql)
	result = cursor.fetchone()
	if result is not None:
		user_id = result[0]
		current_datetime = datetime.now()
		date_string = current_datetime.strftime("%Y-%m-%d %H:%M:%S")
		
		sql = "INSERT INTO `User_Log` ( `time`, `User_id`, `Is_Administrator`, `operate`) VALUES ( '%s', '%s', '%s', '%s');"%(date_string,user_id,str(is_administrator),way)
		cursor.execute(sql)
		db.commit()
		cursor.close()
		db.close()
		return 1
	else:
		cursor.close()
		db.close()
		return 0

@csrf_exempt
def	backend_Get_user(request) :
	if 'backend_login' not in request.session:
		raise PermissionDenied('無效的請求方法')
	with open('static/media/link.json', 'r', encoding='utf-16') as file:
		json_data = json.load(file)
	db = pymysql.connect(host=json_data['db']['host'],
					user=json_data['db']['user'],
					password=json_data['db']['passwd'],
					database=json_data['db']['db'])
	cursor = db.cursor()
	sql = """
	SELECT `Member`.`id`, `Member`.`institution`, `Member`.`account`, `Member`.`name`, `Member`.`email`, `Member`.`phone`, MAX(`Member_Log`.`end_time`) AS `max_end_time`
	FROM `Member`
	LEFT JOIN `Member_Log` ON `Member_Log`.`Member_id` = `Member`.`id`
	GROUP BY `Member`.`id`, `Member`.`institution`, `Member`.`account`, `Member`.`name`, `Member`.`email`, `Member`.`phone`;
	"""
	cursor.execute(sql)
	results = cursor.fetchall()
	data = []

	for row in results:
		data.append({
			'id': row[0],
			'institution': row[1],
			'account':  row[2],
			'name': row[3],
			'email': row[4],
			'phone': row[5],
			'end_time': row[6],
		})
	return JsonResponse(data, safe=False)

@csrf_exempt
def	backend_Modify_user(request) :
	if 'backend_login' not in request.session:
		raise PermissionDenied('無效的請求方法')
	result = {'status':'','msg':''}
	with open('static/media/link.json', 'r', encoding='utf-16') as file:
		json_data = json.load(file)
	db = pymysql.connect(host=json_data['db']['host'],
                     user=json_data['db']['user'],
                     password=json_data['db']['passwd'],
                     database=json_data['db']['db'])
	cursor = db.cursor()

	try:
		id = request.POST['id']
		email = request.POST['email']
		name = request.POST['name']
		account = request.POST['account']
		phone = request.POST['phone']
		institution = request.POST['institution']
		EffectiveDate = request.POST['EffectiveDate']
		ExpiryDate = request.POST['ExpiryDate']
		current_datetime = datetime.now()
		date_string = current_datetime.strftime("%Y-%m-%d %H:%M:%S")
		if check_illegal_parameter(account) or check_illegal_parameter(email) or check_illegal_parameter(name) or check_illegal_parameter(phone) or check_illegal_parameter(institution)  :
			result['status'] = 'No'
			result['msg'] = 'illegal_parameter'

		
		else:
			sql = "UPDATE `Member` SET `account`='%s',  `name`='%s', `email`='%s', `phone`='%s', `institution`='%s', `Log_time`='%s' WHERE `id`=%s"%(account, name, email, phone, institution, date_string,id)
			cursor.execute(sql)
			db.commit()
			
			sql = 'SELECT MAX(`end_time`) AS max_end_time FROM `Member_Log` WHERE `Member_id` = %s;'%id
			cursor.execute(sql)
			sqlres = cursor.fetchone()
			max_end_time = sqlres[0]
			date2 = datetime.strptime(EffectiveDate, '%Y/%m/%d').date()
			if(date2 >max_end_time ):
				sql = "INSERT INTO `Member_Log`(`Member_id`, `start_time`, `end_time`, `Log_time`) VALUES ('%s','%s','%s','%s')"%(id,EffectiveDate,ExpiryDate,date_string)
				cursor.execute(sql)
				db.commit()
			
			result['status'] = 'Yes'

	except Exception as e:
		result['status'] = 'Error'
		result['msg'] = str(e)
	db.close()
	record("編輯會員",request.session['backend_login'],1)
	return JsonResponse(result, safe=False)
@csrf_exempt
def backend_Delete_user(request):
	if 'backend_login' not in request.session:
		raise PermissionDenied('無效的請求方法')
	result = {'status':'','msg':'','data':{}}
	with open('static/media/link.json', 'r', encoding='utf-16') as file:
		json_data = json.load(file)
	db = pymysql.connect(host=json_data['db']['host'],
                     user=json_data['db']['user'],
                     password=json_data['db']['passwd'],
                     database=json_data['db']['db'])
	cursor = db.cursor()
	try:
		id = request.POST['id']
		cursor.execute("DELETE FROM `Member_Log` WHERE Member_id = '%s'"%id)
		cursor.execute("DELETE FROM `Member` WHERE id = '%s'"%id)
		db.commit()
		result['status'] = 'Yes'
	except Exception as e:
		result['status'] = 'Error'
		result['msg'] = str(e)
	cursor.close()
	db.close()
	record("刪除會員",request.session['backend_login'],1)
	return JsonResponse(result, safe=False)
@csrf_exempt
def backend_Notice_user(request):
	if 'backend_login' not in request.session:
		raise PermissionDenied('無效的請求方法')
	result = {'status':'','msg':'','data':{}}
	with open('static/media/link.json', 'r', encoding='utf-16') as file:
		json_data = json.load(file)
	db = pymysql.connect(host=json_data['db']['host'],
                     user=json_data['db']['user'],
                     password=json_data['db']['passwd'],
                     database=json_data['db']['db'])
	cursor = db.cursor()
	try:

		id = request.POST['id']

		sql = "SELECT `email`,`account` FROM `Member` WHERE `id`= '%s';"%id
		cursor.execute(sql)
		sqlres = cursor.fetchone()

		email = sqlres[0]
		account = sqlres[1]
		characters = string.ascii_letters + string.digits
		# 生成 8 碼的隨機亂碼
		password = ''.join(random.choice(characters) for _ in range(8))
		current_datetime = datetime.now()
		date_string = current_datetime.strftime("%Y-%m-%d %H:%M:%S")
		sql = "UPDATE `Member` SET  `pwd`='%s', `Log_time`='%s' WHERE `id`=%s"%(password, date_string,id)
		cursor.execute(sql)
		db.commit()

		subject = "語料庫帳號密碼"
		body = "您的語料庫資訊 \n帳號 : %s  \n密碼 : %s "%(account,password)
		recipients = ["%s"%email]
		send_email(subject, body, recipients)
		result['status'] = 'Yes'
	except Exception as e:
		result['status'] = 'Error'
		result['msg'] = str(e)
	cursor.close()
	db.close()
	record("通知會員帳密",request.session['backend_login'],1)
	return JsonResponse(result, safe=False)

@csrf_exempt
def backend_get_userdata(request) :
	if 'backend_login' not in request.session:
		raise PermissionDenied('無效的請求方法')
	result = {}
	with open('static/media/link.json', 'r', encoding='utf-16') as file:
		json_data = json.load(file)
	db = pymysql.connect(host=json_data['db']['host'],
					user=json_data['db']['user'],
					password=json_data['db']['passwd'],
					database=json_data['db']['db'])
	cursor = db.cursor()
	try:
		id = request.POST['id']
		
		sql = "SELECT `name`,`email`,`phone`,`institution`,`account` FROM `Member` WHERE `id` = '%s';"%id
		cursor.execute(sql)
		sqlres = cursor.fetchone()
		result['name'] = sqlres[0]
		result['email'] = sqlres[1]
		result['phone'] = sqlres[2]
		result['institution'] = sqlres[3]
		result['account'] = sqlres[4]
		result['status'] = 'Yes'

	except Exception as e:
		result['status'] = 'Error'
		result['msg'] = str(e)
	db.close()
	return JsonResponse(result, safe=False)

@csrf_exempt
def	backend_get_user_license(request) :
	if 'backend_login' not in request.session:
		raise PermissionDenied('無效的請求方法')
	with open('static/media/link.json', 'r', encoding='utf-16') as file:
		json_data = json.load(file)
	db = pymysql.connect(host=json_data['db']['host'],
					user=json_data['db']['user'],
					password=json_data['db']['passwd'],
					database=json_data['db']['db'])
	cursor = db.cursor()
	id = request.POST['id']
	sql = "SELECT `start_time`,`end_time` FROM `Member_Log`WHERE `Member_id`='%s'ORDER BY`end_time` ASC;"%id
	cursor.execute(sql)
	results = cursor.fetchall()
	data = []
	for row in results:
		data.append({
			'start': row[0],
			'end':  row[1],
		})
	return JsonResponse(data, safe=False)

@csrf_exempt
def backend_update_speaker(request):
	if 'backend_login' not in request.session:
		raise PermissionDenied('無效的請求方法')
		# 創建資料庫連接
	with open('static/media/link.json', 'r', encoding='utf-16') as file:
		json_data = json.load(file)
	db = pymysql.connect(host=json_data['db']['host'],
					user=json_data['db']['user'],
					password=json_data['db']['passwd'],
					database=json_data['db']['db'])
	try:
		current_datetime = datetime.now()
		date_string = current_datetime.strftime("%Y-%m-%d %H:%M:%S")
		cursor = db.cursor()
		sql = "SELECT * FROM `Speaker_data` "
		cursor.execute(sql)
		has_data = cursor.fetchall()
		if (has_data) :
			db.close()
			return JsonResponse({'message': '資料表已有資料，請先刪除在上傳'})
		speaker_file = request.FILES.get('speaker_file')

		# 打開文件
		xlsx = pd.ExcelFile(speaker_file)

		# 獲取所有工作表名稱
		sheet_names = xlsx.sheet_names


		# 寫入到資料庫中
		for sheet_name in sheet_names:
			theme = sheet_name
			df1 = pd.read_excel(speaker_file, sheet_name=theme )
			df1.fillna(value='未知', inplace=True)
			if(theme == 'TMC') :
				data = []
				for index, row in df1.iterrows():
					data.append((row['對話編號(原始)'],row['聲道(原始)'],row['聲道編號'],row['對話編號'],row['語者編號'],row['性別'],row['年齡'],row['錄音地點'],date_string,theme))
				sql = "INSERT INTO Speaker_data(Talk_id_original, sound_track_original, sound_track, Talk_id, Speaker_id, sex, age, location, Log_time,Classification) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s,%s)"  
				cursor.executemany(sql,data)
				db.commit()
			if(theme== 'Socio') :
				data = []
				for index, row in df1.iterrows():
					data.append((row['語者編號'],row['性別'],row['年齡'],row['錄音地點'],date_string,theme))          

				sql = "INSERT INTO Speaker_data(Speaker_id, sex, age, location, Log_time,Classification) VALUES ( %s, %s, %s, %s, %s,%s)" 
				cursor.executemany(sql,data)
				db.commit()

			if(theme== 'Child_HI'or theme=='Child_NH') :
				data = []
				for index, row in df1.iterrows():
					data.append((row['語者編號(原始)'],row['語者編號'],row['性別'],row['年齡'],row['錄音地點'],date_string,theme))
				sql = "INSERT INTO Speaker_data(Speaker_id_original,Speaker_id, sex, age, location, Log_time,Classification) VALUES ( %s, %s, %s, %s, %s,%s,%s)"  
				cursor.executemany(sql,data)
				db.commit()

		cursor.close()
		db.close()
		record("上傳語者檔",request.session['backend_login'],1)
		return JsonResponse({'message': '新增成功'})
	except Exception as e:
		db.close()
		return JsonResponse({'message': '上傳檔案錯誤，請確認檔案格式'})
@csrf_exempt
def backend_download_speaker(request):
	if 'backend_login' not in request.session:
		raise PermissionDenied('無效的請求方法')
		# 連接資料庫
	with open('static/media/link.json', 'r', encoding='utf-16') as file:
		json_data = json.load(file)
	db = pymysql.connect(host=json_data['db']['host'],
					user=json_data['db']['user'],
					password=json_data['db']['passwd'],
					database=json_data['db']['db'])
	cursor = db.cursor()
	sql = "SELECT * FROM `Speaker_data` "
	cursor.execute(sql)
	has_data = cursor.fetchall()
	if not (has_data) :
		db.close()
		return JsonResponse({'message': '資料表遺失，請重新上傳'})
	# 創建一個新的工作簿
	workbook = openpyxl.Workbook()
	from openpyxl.styles import Alignment
	# 刪除預設的工作表
	default_sheet = workbook['Sheet']
	workbook.remove(default_sheet)

	alignment = Alignment(horizontal='center', vertical='center')

	worksheet = workbook.create_sheet(title='TMC')
	# 寫入標題行
	headers = ['對話編號(原始)', '聲道(原始)', '聲道編號', '對話編號', '語者編號', '性別', '年齡', '錄音地點', '', '聲道(原始)', '聲道']
	# 設定每個列的寬度
	column_widths = {'A': 16.5, 'B': 10.17, 'C': 8.5, 'D': 13, 'E': 15.67, 'F': 6, 'G': 7.17,'J':11}  

	for col, header in enumerate(headers, start=1):
		column_letter = openpyxl.utils.get_column_letter(col)
		worksheet.cell(row=1, column=col, value=header)
		if column_letter in column_widths:
				worksheet.column_dimensions[column_letter].width = column_widths[column_letter]
	font = Font(name='微軟正黑體' ,bold=True)
	font2 = Font(name='微軟正黑體' ,bold=True,color="BFBFBF")
	fill = PatternFill(fill_type='solid', fgColor='D9D9D9')
	for cell in worksheet[1]:
		if(cell.column_letter =='A' or cell.column_letter =='B' or cell.column_letter =='C'  ):
			cell.font = font2
		else :
			cell.font = font
		cell.alignment = alignment
		cell.fill = fill

	sql = "SELECT `Talk_id_original`,`sound_track_original`,`sound_track`,`Talk_id`,`Speaker_id`,`sex`,`age`,`location` FROM `Speaker_data` WHERE `Classification` = 'TMC';"
	cursor.execute(sql)
	results = cursor.fetchall()
	fill = PatternFill(fill_type='solid', fgColor='EDEDED')
	font = Font(name='微軟正黑體')
	font2 = Font(name='微軟正黑體' ,color="BFBFBF")
	for row in results:
		data = [row[0],row[1], row[2], row[3], row[4], row[5], row[6], row[7]]
		worksheet.append(data)
		last_row = worksheet.max_row

		if(row[2]== 'A') :
			for row in worksheet.iter_rows(min_row=last_row, max_row=last_row):
				for cell in row:
					if(cell.column_letter== "I" or cell.column_letter== "J" or cell.column_letter== "K"  ) :
						continue
					cell.fill = fill
	worksheet['J2'] = '左'
	worksheet['k2'] = 'A'
	worksheet['J2'].fill = fill
	worksheet['k2'].fill = fill
	worksheet['J3'] = '右'
	worksheet['k3'] = 'B'
	font = Font(name='微軟正黑體')
	font2 = Font(name='微軟正黑體' ,color="BFBFBF")
	for row in worksheet.iter_rows(min_row=2):  # 從第二列開始
		for cell in row:
			if(cell.column_letter =='A' or cell.column_letter =='B' or cell.column_letter =='C'  ):
				cell.font = font2
			else :
				cell.font = font
			cell.alignment = alignment
	#寫入Socio
	worksheet = workbook.create_sheet(title='Socio')
	# 寫入標題行
	headers = [ '語者編號', '性別', '年齡', '錄音地點']
	# 設定每個列的寬度
	column_widths = {'A': 10.17, 'B': 10.67, 'C': 9, 'D': 13} 
	for col, header in enumerate(headers, start=1):
		column_letter = openpyxl.utils.get_column_letter(col)
		worksheet.cell(row=1, column=col, value=header)
		if column_letter in column_widths:
				worksheet.column_dimensions[column_letter].width = column_widths[column_letter]
	font = Font(name='微軟正黑體' ,bold=True)
	fill = PatternFill(fill_type='solid', fgColor='D9D9D9')
	for cell in worksheet[1]:
		cell.alignment = alignment
		cell.font = font
		cell.fill = fill
	sql = "SELECT `Speaker_id`,`sex`,`age`,`location` FROM `Speaker_data` WHERE `Classification` = 'Socio';"
	cursor.execute(sql)
	results = cursor.fetchall()
	font = Font(name='微軟正黑體')
	for row in results:
		data = [row[0],row[1], row[2], row[3]]
		worksheet.append(data)
		last_row = worksheet.max_row
		for row in worksheet.iter_rows(min_row=last_row, max_row=last_row):
			for cell in row:
				cell.font = font
				cell.alignment = alignment
	#寫入Child
	child_list=['Child_HI','Child_NH']
	for each in child_list :
		worksheet = workbook.create_sheet(title=each)
			# 寫入標題行
		headers = [ '語者編號(原始)','語者編號', '性別', '年齡', '錄音地點']
		# 設定每個列的寬度
		column_widths = {'A': 15.17, 'B': 15.67} 
		for col, header in enumerate(headers, start=1):
			column_letter = openpyxl.utils.get_column_letter(col)
			worksheet.cell(row=1, column=col, value=header)
			if column_letter in column_widths:
					worksheet.column_dimensions[column_letter].width = column_widths[column_letter]
		
		font = Font(name='微軟正黑體' ,bold=True)
		font2 = Font(name='微軟正黑體' ,bold=True,color="BFBFBF")
		fill = PatternFill(fill_type='solid', fgColor='D9D9D9')
		for cell in worksheet[1]:
			if(cell.column_letter =='A'):
				cell.font = font2
			else :
				cell.font = font
			cell.alignment = alignment
			cell.fill = fill
		
		sql = "SELECT  `Speaker_id_original`, `Speaker_id`,`sex`,`age`,`location` FROM `Speaker_data` WHERE `Classification` = '%s';"%each
		cursor.execute(sql)
		results = cursor.fetchall()
		font = Font(name='微軟正黑體')
		font2 = Font(name='微軟正黑體' ,color="BFBFBF")
		for row in results:
			data = [row[0],row[1], row[2], row[3], row[4]]
			worksheet.append(data)
			last_row = worksheet.max_row
			for row in worksheet.iter_rows(min_row=last_row, max_row=last_row):
				for cell in row:
					if(cell.column_letter =='A'):
						cell.font = font2 
					else :
						cell.font = font
					cell.alignment = alignment


	    # 設定下載檔案的檔名	
    
	file_name = 'Sinca_Corpora_Speaker.xlsx'
	# 將Excel工作簿轉換為二進位數據

	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	
	response['Content-Disposition'] = 'attachment; filename="{}"'.format(file_name)

	record("下載語者檔",request.session['backend_login'],1)
	workbook.save(response)
	db.close()
	return response

@csrf_exempt
def backend_delete_speaker(request) :
	if 'backend_login' not in request.session:
		raise PermissionDenied('無效的請求方法')

	with open('static/media/link.json', 'r', encoding='utf-16') as file:
		json_data = json.load(file)
	db = pymysql.connect(host=json_data['db']['host'],
                     user=json_data['db']['user'],
                     password=json_data['db']['passwd'],
                     database=json_data['db']['db'])

	cursor = db.cursor()
	sql = "SELECT * FROM `Speaker_data` "
	cursor.execute(sql)
	has_data = cursor.fetchall()
	if not (has_data) :
		db.close()
		return JsonResponse({'message': '資料表以為空'})

	cursor.execute("DELETE FROM `Speaker_data`")
	db.commit()
	cursor.close()
	db.close()
	record("刪除語者檔",request.session['backend_login'],1)
	return JsonResponse({'message': '刪除成功'})

@csrf_exempt
def backend_new_resource (request) :
	if 'backend_login' not in request.session:
		raise PermissionDenied('無效的請求方法')
	

	uploaded_file = request.FILES['file']
	resource_name = request.POST.get('resource_name')
	resource_content = request.POST.get('resource_content')
	current_datetime = datetime.now()
	date_string = current_datetime.strftime("%Y-%m-%d %H:%M:%S")
	
	with open('static/media/link.json', 'r', encoding='utf-16') as file:
		json_data = json.load(file)
	db = pymysql.connect(host=json_data['db']['host'],
					user=json_data['db']['user'],
					password=json_data['db']['passwd'],
					database=json_data['db']['db'])
	cursor = db.cursor()
	sql = "SELECT * FROM `Authorized_File` WHERE `download_name`='%s'"%uploaded_file.name
	cursor.execute(sql)
	cursor.execute(sql)
	result = cursor.fetchone()
	if(result) :
		cursor.close()
		db.close()
		return JsonResponse({'message': '已存在相同檔案'})

	sql = "INSERT INTO `Authorized_File`( `FName`, `Fcontent`, `download_name`, `Log_time`) VALUES ('%s','%s','%s','%s')"%(resource_name,resource_content,uploaded_file.name,date_string)
	cursor.execute(sql)
	db.commit()
	cursor.close()
	db.close()
	# 在伺服器上儲存檔案
	with open( 'static/upload/'+uploaded_file.name, 'wb+') as destination:
		for chunk in uploaded_file.chunks():
			destination.write(chunk)
	record("新增授權資源",request.session['backend_login'],1)
	return JsonResponse({'message': '新增成功！'})
@csrf_exempt
def backend_get_resource (request) :
	with open('static/media/link.json', 'r', encoding='utf-16') as file:
		json_data = json.load(file)
	db = pymysql.connect(host=json_data['db']['host'],
					user=json_data['db']['user'],
					password=json_data['db']['passwd'],
					database=json_data['db']['db'])
	cursor = db.cursor()
	sql = "SELECT `id`,`FName`,`Fcontent`,`download_name`,`time` FROM `Authorized_File`;"
	cursor.execute(sql)
	results = cursor.fetchall()
	data = []
	for row in results:
		data.append({
			'id': row[0],
			'FName':  row[1],
			'Fcontent':  row[2],
			'download_name':  row[3],
			'time':  row[4],
		})
	return JsonResponse(data, safe=False)

@csrf_exempt
def backend_delete_resource (request) :	
	if 'backend_login' not in request.session:
		raise PermissionDenied('無效的請求方法')
	id = request.POST.get('id')
	with open('static/media/link.json', 'r', encoding='utf-16') as file:
		json_data = json.load(file)
	db = pymysql.connect(host=json_data['db']['host'],
					user=json_data['db']['user'],
					password=json_data['db']['passwd'],
					database=json_data['db']['db'])
	cursor = db.cursor()
	sql = "SELECT `download_name` FROM `Authorized_File` WHERE `id`='%s';"%id
	cursor.execute(sql)
	results = cursor.fetchone()
	file_path = 'static/upload/'+results[0] # 指定要刪除的檔案路徑

	if os.path.exists(file_path):
		os.remove(file_path)
		sql = "DELETE FROM `Authorized_File` WHERE  `id`='%s';"%id
		cursor.execute(sql)
		db.commit()
		db.close()
		record("刪除授權資源",request.session['backend_login'],1)
		return JsonResponse({'message': '刪除成功'})
	else:
		db.close()
		return JsonResponse({'message': '檔案不存在於伺服器'})



@csrf_exempt
def download_resource (request) :
	with open('static/media/link.json', 'r', encoding='utf-16') as file:
		json_data = json.load(file)
	db = pymysql.connect(host=json_data['db']['host'],
					user=json_data['db']['user'],
					password=json_data['db']['passwd'],
					database=json_data['db']['db'])
	cursor = db.cursor()
	file_name = request.POST.get('file_name')
	sql = "SELECT `time` FROM `Authorized_File` WHERE `download_name`='%s'"%file_name
	cursor.execute(sql)
	result = cursor.fetchone()
	if not (result) :
		cursor.close()
		db.close()
		return JsonResponse({'message': '下載錯誤 : 檔案不存在'})
	time = result[0]
	time = int(time)
	time+=1
	current_datetime = datetime.now()
	date_string = current_datetime.strftime("%Y-%m-%d %H:%M:%S")
	sql = "UPDATE `Authorized_File` SET `time`='%s',`Log_time`='%s' WHERE `download_name` ='%s'"%(time,date_string,file_name)
	cursor.execute(sql)
	db.commit()
	file_path = 'static/upload/'+ file_name # 指定下載的檔案路徑
	with open(file_path, 'rb') as f:
		response = HttpResponse(f.read(), content_type='application/octet-stream')
		response['Content-Disposition'] = f'attachment; filename="{file_name}"'
		return response

def send_email( subject, body, recipients, attachments=None):
    try:
        # 設定郵件內容
        msg = MIMEMultipart()
        msg['From'] = 'lingmis@gate.sinica.edu.tw'
        msg['To'] = ", ".join(recipients)
        msg['Subject'] = subject

        # 添加郵件內容
        msg.attach(MIMEText(body, 'plain'))

        if attachments:
            for attachment in attachments:
                with open(attachment, "rb") as file:
                    part = MIMEApplication(file.read())
                    part.add_header('Content-Disposition', 'attachment', filename=attachment)
                    msg.attach(part)

        # 連接到SMTP伺服器並寄送郵件
        smtp_server = 'smtp.sinica.edu.tw' 
        smtp_port = 25
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login('lingmis@gate.sinica.edu.tw', 'ling0201')
            server.send_message(msg)
        print("郵件已成功寄出！")
    except Exception as e:
        print("郵件寄送失敗：", str(e))

@csrf_exempt
def download_pdf(request):
    # 獲取要下載的文件路徑
    file_path = os.path.join(settings.STATIC_ROOT, 'pdf', '中研院中文口語語料庫檢索系統使用書.pdf')
    
     # 檢查文件是否存在
    if os.path.exists(file_path):
        with open(file_path, 'rb') as file:
            # 創建響應對象，將文件内容作為響應的内容
            response = HttpResponse(file.read(), content_type='application/pdf')

            # 設置文件的下載頭訊息
            response['Content-Disposition'] = 'attachment;'

            return response

    # # 文件不存在時顯示錯誤
    return HttpResponse('File not found.')
@csrf_exempt
def send_password(request) :
	result = {'status':'','msg':''}
	with open('static/media/link.json', 'r', encoding='utf-16') as file:
		json_data = json.load(file)
	db = pymysql.connect(host=json_data['db']['host'],
                     user=json_data['db']['user'],
                     password=json_data['db']['passwd'],
                     database=json_data['db']['db'])
	cursor = db.cursor()
	try:
		email = request.POST['email']
		if check_illegal_parameter(email):
			result['status'] = 'illegal'
			result['msg'] = '不合法字元'
		else:
			# 尋找輸入的email使否有與之相符的會員
			sql = "SELECT * FROM `Member` WHERE `email`='%s';"%email
			cursor.execute(sql)
			has_email = cursor.fetchall()
			if has_email :
				result['status'] = 'Yes'
				# 產生10碼由數字及字母組成的隨機密碼
				characters = string.ascii_letters + string.digits
				password = ''.join(random.choice(characters) for _ in range(10))
				subject = "中研院中文口語語料庫重置密碼"
				body = "您的語料庫資訊 \n帳號 : %s  \n新密碼 : %s "%(email,password)
				recipients = ["%s"%email]
				send_email(subject, body, recipients)
				# 將該會員原本的密碼更改為產生的隨機密碼
				sql = "UPDATE `Member` SET `pwd` = '%s' WHERE `Member`.`email` = '%s'"%(password,email)
				cursor.execute(sql)
				db.commit()
				result['msg'] = '傳送成功'
			else :
				result['status'] = 'No'
				result['msg'] = '查無此信箱,請輸入申請帳號時使用的信箱!'
	except Exception as e:
		result['status'] = 'Error'
		result['msg'] = str(e)
	db.close()
	return JsonResponse(result, safe=False)

@csrf_exempt
def upload_application (request) :
	uploaded_file = request.FILES['file']
	# 在伺服器上儲存檔案
	with open( 'static/application/'+uploaded_file.name, 'wb+') as destination:
		for chunk in uploaded_file.chunks():
			destination.write(chunk)
	return JsonResponse({'message': '上傳成功！'})
    
@csrf_exempt
def apply_confirm(request) :
	uploaded_file = request.FILES['file']
	STATIC_URL = '/static/'
	file_path = os.path.join(settings.BASE_DIR, 'static', 'application', uploaded_file.name)
	result = {'status':'','msg':''}
	with open('static/media/link.json', 'r', encoding='utf-16') as file:
		json_data = json.load(file)
	db = pymysql.connect(host=json_data['db']['host'],
                     user=json_data['db']['user'],
                     password=json_data['db']['passwd'],
                     database=json_data['db']['db'])
	cursor = db.cursor()
	try:
		# 找到所有管理員的email
		sql = "SELECT `email` FROM `Administrator`"
		cursor.execute(sql)
		result['status'] = 'Yes'
		subject = "中研院中文口語語料庫會員申請"
		body = "附件是會員申請書"
		# 將收件人設定為所有的管理員
		recipients = [row[0] for row in cursor.fetchall()]
		# 將申請書以信件附件方式傳送
		attachments=[file_path]
		send_email(subject, body, recipients,attachments)
		db.commit()
		return JsonResponse({'message': '寄送成功！'})
	except Exception as e:
		result['status'] = 'Error'
		result['msg'] = str(e)
	db.close()
	return JsonResponse(result, safe=False)

@csrf_exempt
def search_corpus(request) :
	condition_json = request.POST['condition']
	selected_option = request.POST.get('selected_value')
	check_options = request.POST.get('checkbox_value')

	selected_option = json.loads(selected_option)
	check_options = json.loads(check_options)
	condition = json.loads(condition_json)
	main_value = request.POST['main_value']
	mode = int(request.POST['mode'])
	temp_pos = 0
	total =0
	with open('static/media/link.json', 'r', encoding='utf-16') as file:
		json_data = json.load(file)
	db = pymysql.connect(host=json_data['db']['host'],
                     user=json_data['db']['user'],
                     password=json_data['db']['passwd'],
                     database=json_data['db']['db'])
	cursor = db.cursor()
	search_pattern = []
	addtion_search = []
	cross_ipu =False
	search_pattern.append(main_value)
	for each in  condition :
		if each[0] == "or" :
			search_pattern.append(each[1])
		else :
			addtion_search.append(each)
	if(mode == 2) :
		temp_pos = int(request.POST['text_number'])
		cross_ipu = True

	temp=[]
	temp2=[]
	"""
	selected_option = 2 社會
	selected_option = 3 兒童
	check_options = ５ 聽常
	check_options = ６ 聽損
	"""

	if selected_option == '2':
		sql = "SELECT `IPUStart`,`IPUEnd`,`IPU`,`Filename` FROM `S_File` ORDER BY `Sid` ASC;"
	else :
		if len(check_options) == 2 :
			sql = "SELECT `IPUStart`,`IPUEnd`,`IPU`,Child.Speaker_id,Child.theme FROM `Child_File` JOIN  `Child`  ON Child_File.Cid = Child.id ;"
		elif '5' in check_options :
			sql = "SELECT `IPUStart`,`IPUEnd`,`IPU`,Child.Speaker_id,Child.theme  FROM `Child_File` JOIN  `Child`  ON Child_File.Cid = Child.id  WHERE Child.theme ='Child_HI';"
		else :
			sql = "SELECT `IPUStart`,`IPUEnd`,`IPU`,Child.Speaker_id,Child.theme  FROM `Child_File` JOIN  `Child`  ON Child_File.Cid = Child.id  WHERE Child.theme ='Child_NH';"
	
	cursor.execute(sql)
	results = cursor.fetchall()
	temp3=[]
	start_id = results[0][3]
	for row in results: 
		total +=1		
		if(selected_option == '2') :
			data = Content()
			data.ipustart = row[0]
			data.ipuend = row[1]
			data.ipu = row[2]
			data.filename = row[3]
			data.theme = "社會語言學問卷語音資料庫"
			if(start_id != row[3]) :
				start_id = row[3]
				temp2.append(temp)
				temp=[]
			temp.append(data)
			
		else : 
			data = Content()
			data.ipustart = row[0]
			data.ipuend = row[1]
			data.ipu = row[2]
			data.filename = row[3]
			if(row[4] == 'Child_HI') :
				data.theme = "說故事─龜兔賽跑（聽常兒童）"
			else:
				data.theme = "說故事─龜兔賽跑（聽損兒童）"
			if(start_id != row[3]) :
				start_id = row[3]
				temp2.append(temp)
				temp=[]
			temp.append(data)
	if(temp) :
		temp2.append(temp)
	
	for data in temp2 :
		find = True
		if(cross_ipu) :
			start_index = 0
			concatenated_string = ''
			string_list = []
			for row in data:
				row.start_pos = start_index
				start_index += len(row.ipu)
				string_list.append(row.ipu) 
			concatenated_string = ''.join(string_list)
			for index, row in enumerate(data):
				has_keyword = False
				pos =0
				for keyword in search_pattern :
					if (keyword in row.ipu) :
						temp3.append(row)
						pos = row.ipu.find(keyword)
						has_keyword = True
						break
				if(has_keyword) :
					for each in addtion_search :
						if(row.start_pos+pos -temp_pos < 0) :
							sub = concatenated_string[0:row.start_pos+pos +temp_pos]
						else :
							sub = concatenated_string[row.start_pos+pos -temp_pos:row.start_pos+pos +temp_pos] 
						if(each[0] == "not"):
							if(each[1] in sub) :
								find = False
								temp3.pop()
								break
						if(each[0] == "and"):
							if(each[1] not in sub) :
								find = False
								temp3.pop()
								break
				if(find) :
					start_time = data[index].ipustart
					temp2 = []
					#抓取前一分鐘的字
					for i in range(index-1 , 0 , -1) :
						temp2.insert(0,data[i].ipu)                  
						if(start_time - data[i].ipustart > 60) :
							break
					#保留初始數據
					temp2.append( data[index].ipu)
					#抓取後一分鐘的字
					for i in range(index+1 , len(data)) :
						temp2.append(data[i].ipu)
						if(data[i].ipustart - start_time >60) :
							break
					row.sub_content_array = temp2
		else :
			for index, row in enumerate(data):
				has_keyword = False 
				for keyword in search_pattern :
					if (keyword in row.ipu) :
						temp3.append(row)
						has_keyword = True
						break
				if(has_keyword) :
					for each in addtion_search :
						if(each[0] == "not"):
							if(each[1] in row.ipu) :
								find = False
								temp3.pop()
								break
							else :
								find = True
							
						if(each[0] == "and"):
							if(each[1] not in row.ipu) :
								find = False
								temp3.pop()
								break
							else :
								find = True     
				if(find) :
					start_time = data[index].ipustart
					temp2 = []
					#抓取前一分鐘的字
					for i in range(index-1 , 0 , -1) :
						temp2.insert(0,data[i].ipu)                  
						if(start_time - data[i].ipustart > 60) :
							break
					#保留初始數據
					temp2.append( data[index].ipu)
					#抓取後一分鐘的字
					for i in range(index+1 , len(data)) :
						temp2.append(data[i].ipu)
						if(data[i].ipustart - start_time >60) :
							break
					row.sub_content_array = temp2
		
	db.close()
	json_data = json.dumps([content.to_json() for content in temp3], indent=4)
	return JsonResponse({'total': total,'serach_num':len(temp3),'search_corpus':json_data})

@csrf_exempt
def search_MHcorpus(request):
	main = json.loads(request.POST['main'])
	pos = 0
	mode = int(request.POST['mode'])
	if(mode==2 or mode==3) :
		pos = int(request.POST['text_number'])
	check = json.loads(request.POST['check'])
	check2=[]
	for each in check :
		if('checkbox2' ==each ) :
			check2.append('MCDC_8')
		if('checkbox3' ==each ) :
			check2.append('MCDC_22')
		if('checkbox1' ==each ) :
			check2.append('MTCC')
		if('checkbox4' ==each ) :
			check2.append('MMTC')
	condition = json.loads(request.POST['condition'])
	with open('static/media/link.json', 'r', encoding='utf-16') as file:
		json_data = json.load(file)
	db = pymysql.connect(host=json_data['db']['host'],
						user=json_data['db']['user'],
						password=json_data['db']['passwd'],
						database=json_data['db']['db'])
	cursor = db.cursor()  
	temp_sql=""
	for each in check2 :
		if(temp_sql == "") :
			temp_sql+="Modern_Chinese.theme = '%s'"%each
		else :
			temp_sql+=" or Modern_Chinese.theme = '%s'"%each
	sql = "SELECT `IPUStart`,`IPUEnd`,`IPU`,`Word`,`POS`,`Pinyin` ,`FileSn`,Modern_Chinese.Speaker_id   ,Modern_Chinese.theme  FROM `MH_File` JOIN  `Modern_Chinese`  ON Modern_Chinese.id = MH_File.MH_id   WHERE %s ORDER BY `FileSn` ASC , `IPUStart` ASC;"%temp_sql
	cursor.execute(sql)
	results = cursor.fetchall()
	temp=[]
	temp2=[]
	theme = ""
	total_count = 0
	Talk_id = results[0][6]
	for each in results :
		total_count+=1
		my_data = Content_conversation()
		my_data.ipu = each[2]
		my_data.word = each[3]
		my_data.start_time = each[0]
		my_data.end_time = each[1]
		my_data.pinyin = each[5]
		theme = each[8]
		my_data.pos = each[4]
		if( "_A" in each[7]) :
			my_data.speaker = "A"
		else :
			my_data.speaker = "B"
		if(each[6] != Talk_id) :
			#變更下一段對話編號
			my_conversation = Conversation()
			my_conversation.content_array = temp
			my_conversation.talk_id = Talk_id
			my_conversation.theme = theme
			Talk_id = each[6]
			temp2.append(my_conversation)
			temp = []
		temp.append(my_data)
	if(temp) :
		my_conversation = Conversation()
		my_conversation.content_array = temp
		my_conversation.theme = theme
		my_conversation.talk_id = Talk_id
		temp2.append(my_conversation)
	search_pattern=[]
	addtion_search=[]
	search_pattern.append(main)
	for each in condition :
		if(each[0] == 'or') :
			search_pattern.append([each[1],each[2]])
		else :
			addtion_search.append(each)
	results=[]
	for data_content in temp2 :
		#單獨ipu
		if(mode == 1) :
			content_data = data_content.content_array

			for index, each in enumerate(content_data):
				is_matched = False
				has_keyword = False
				for keyword in search_pattern :
					if (keyword[0] in each.ipu or keyword[0] in each.pinyin) :
						word = each.word.split(" ")
						verb = each.pos.split(" ")
						is_matched = True
						has_keyword = True 
						if(is_chinese_string(keyword[0]) and keyword[1] != "") :
							try :
								my_index = word.index(keyword[0])
								if(verb[my_index] == keyword[1]) :
									is_matched = True
									has_keyword = True 
								else :
									is_matched = False
									has_keyword = False 									
							except :
								is_matched = False
								has_keyword = False 
						elif(is_chinese_string(keyword[0])==False and keyword[1] != ""):
							word = each.word.split(" ")
							pinyin_word = pinyin2word(word,each.pinyin)
							try :
								my_index = pinyin_word.index(keyword[0])
								if(verb[my_index] == keyword[1]) :
									is_matched = True
									has_keyword = True 
								else :
									is_matched = False
									has_keyword = False 			
							except :
								is_matched = False
								has_keyword = False 							

					if(has_keyword) :
						for add in addtion_search :	
							if('\u4e00' <= add[1] <= '\u9fff') :																
								if(add[0] == "not"):
									if(add[1] in each.ipu ) :
										is_matched = False
										break
								if(add[0] == "and"):									
									if(add[1] not in each.ipu) :
										is_matched = False
										break
									if( add[2] != "") :
										word = each.word.split(" ")
										verb = each.pos.split(" ")
										try :
											my_index = word.index(add[1])
											if(verb[my_index] != add[2]) :
												is_matched = False							
										except :
											is_matched = False
							else :
								if(add[0] == "not"):
									if(add[1] in each.pinyin) :
										is_matched = False
										break
								if(add[0] == "and"):
									if( add[1] not in each.pinyin) :
											is_matched = False
											break
									if(add[2]!="") :
										word = each.word.split(" ")
										pinyin_word = pinyin2word(word,each.pinyin)
										try :
											
											my_index = pinyin_word.index(add[1])
											if(verb[my_index] != add[2]) :
												is_matched = False		
										except :
											is_matched = False
				if(is_matched) :
					text_count = 20
					start_time = content_data[index].start_time
					temp = []
					temp2 = []
					#抓取前20個字
					for i in range(index-1 , 0 , -1) :
						#保持在第一位
						temp.insert(0,content_data[i])
						if(text_count - len(content_data[i].ipu) < 0) :
							break
						text_count = text_count - len(content_data[i].ipu)
					#抓取前一分鐘的字
					for i in range(index-1 , 0 , -1) :
						temp2.insert(0,content_data[i])                  
						if(start_time - content_data[i].start_time > 60) :
							break
					#保留初始數據
					temp.append(content_data[index])
					temp2.append(content_data[index])
					text_count = 20
					#抓取後一分鐘的字
					for i in range(index+1 , len(content_data)) :
						temp2.append(content_data[i])
						if(content_data[i].start_time - start_time >60) :
							break
					#抓取後20個字
					for i in range(index+1 , len(content_data)) :
						temp.append(content_data[i])
						if(text_count - len(content_data[i].ipu) < 0) :
							break
						text_count = text_count - len(content_data[i].ipu)
					my_conversation = Conversation()
					my_conversation.content_array = temp
					my_conversation.sub_content_array = temp2
					my_conversation.talk_id = data_content.talk_id
					my_conversation.theme = data_content.theme
					results.append(my_conversation)
		#跨ipu 相同語者 & 跨ipu 不同語者
		elif(mode == 2 or mode==3) :
			content_data = data_content.content_array
			for index, row in enumerate(content_data):
				has_keyword = False
				is_matched = False
				find_pos= False
				my_keyword = ""
				for keyword in search_pattern :      
					if (keyword[0] in row.ipu or keyword[0] in row.pinyin) :
						has_keyword = True
						is_matched = True
						if(is_chinese_string(keyword[0]) and keyword[1] != "") :

							word = row.word.split(" ")
							verb = row.pos.split(" ")
							try :
								my_index = word.index(keyword[0])
								if(verb[my_index] == keyword[1]) :
									is_matched = True
									has_keyword = True 
									find_pos = True
								else :
									is_matched = False
									has_keyword = False 									
							except :
								is_matched = False
								has_keyword = False 
						elif(is_chinese_string(keyword[0])==False and keyword[1] != ""):
							verb = row.pos.split(" ")
							word = row.word.split(" ")
							pinyin_word = pinyin2word(word,row.pinyin)
							try :
								my_index = pinyin_word.index(keyword[0])
								if(verb[my_index] == keyword[1]) :
									is_matched = True
									has_keyword = True 
									find_pos = True
								else :
									is_matched = False
									has_keyword = False 			
							except :
								is_matched = False
								has_keyword = False 
						my_keyword = keyword[0]
						if(has_keyword):
							break  
				if(has_keyword) :
					if(is_chinese_string(my_keyword)) :
						#抓取前n個字
						text_count = pos
						text_count2 = pos
						sub = ""
						sub2 = ""
						front = row.ipu[0:row.ipu.find(my_keyword)]
						keyword_pinyin_pos = get_chinese_index(row.ipu.find(my_keyword),row.ipu)
						pinyin = row.pinyin.split(" ")
						pinyin = pinyin[0:keyword_pinyin_pos]
						if(find_pos):
							front_word = row.word[0:row.word.find(my_keyword)].split(" ")							
							front_word = merge_continuous_english_words(front_word)
							post_word = row.word[row.word.find(my_keyword)+len(my_keyword):len(row.word)].split(" ")
							post_word = merge_continuous_english_words(post_word)	
							front_pos = row.pos.split(" ")
							front_pos2 = row.pos.split(" ")						
							if(front_word):
								if(front_word[len(front_word) -1])=="" :
									front_word.pop()
							my_index = len(front_word)-1
							my_index2 = len(front_word)
							post_pos = row.pos.split(" ")
							post_pos = post_pos[my_index2+1:len(post_pos)]
							my_word = []							
							my_pos=[]
							temp_word=""						
						for i in range(len(pinyin)-1,-1,-1) :
								if(text_count2!=0) :
									sub2= pinyin[i] +" "+sub2
									if(contain_digital(pinyin[i])) :
										text_count2-=1  		 
						for i in range(len(front)-1,-1 ,-1) :
							if(text_count!=0) :
								sub = front[i] + sub
								if(find_pos) :
									temp_word=front[i] + temp_word
									if(front_word[my_index] in temp_word) :
										my_word.insert(0,front_word[my_index])
										my_pos.insert(0,front_pos[my_index])
										my_index-=1
										temp_word=""
								if(is_chinese_string(front[i])) :
									text_count-=1					
						for i in range(index-1 , -1 , -1) :
							if(text_count2 <= 0) :
								break
							if(mode==2) :
								if(content_data[i].speaker != row.speaker):
									continue
							pinyin = content_data[i].pinyin.split(" ")
							for q in range(len(pinyin)-1,-1 ,-1) :
								if(text_count2!=0) :
									sub2= pinyin[q] +" "+sub2
									if(contain_digital(pinyin[q])) :
										text_count2-=1                        
						for i in range(index-1 , -1 , -1) :

							if(text_count == 0) :
								break
							if(mode==2) :
								if(content_data[i].speaker != row.speaker):
									continue
							if(find_pos) :							
								front_word = content_data[i].word.split(" ")
								front_word = merge_continuous_english_words(front_word)
								front_pos = content_data[i].pos.split(" ")
								temp_word=""
								my_index = len(front_word)-1
							for q in range(len(content_data[i].ipu)-1,-1 ,-1) :

								if(text_count!=0) :
									if(find_pos) :
										temp_word= content_data[i].ipu[q] + temp_word
										if(front_word[my_index] in temp_word) :
											my_word.insert(0,front_word[my_index])
											my_pos.insert(0,front_pos[my_index])
											my_index-=1
											temp_word=""
									sub = content_data[i].ipu[q] + sub
									if(is_chinese_string(content_data[i].ipu[q])) :
										text_count-=1
						

						#補上自己
						if(find_pos) :
							my_word.append(my_keyword)
							my_pos.append(front_pos2[my_index2])							
						sub+= my_keyword
						pinyin = row.pinyin.split(" ")
						try :
							for i in range(keyword_pinyin_pos , keyword_pinyin_pos+len(keyword)) :
								sub2 += pinyin[i] +" "
						except :
							print('Nuknow_error')
						#抓取後n個字


						post2 = pinyin[keyword_pinyin_pos+len(my_keyword) : len(pinyin)]
						post = row.ipu[row.ipu.find(my_keyword)+len(my_keyword):len(row.ipu)]
						text_count = pos
						text_count2 = pos
						temp_word = ""
						my_index=0
						for i in range(0,len(post)) :
							if(text_count!=0) :								
								sub += post[i] 
								if(find_pos) :
									temp_word+=post[i]
									if(post_word[my_index] in temp_word) :
										my_word.append(post_word[my_index])
										my_pos.append(post_pos[my_index])
										my_index+=1
										temp_word=""

								if(is_chinese_string(post[i])) :
									text_count-=1
						for i in range(0,len(post2)) :
							if(text_count2!=0) :
								sub2 = sub2 +post2[i] +" "
								if(contain_digital(post2[i])) :
									text_count2-=1
							else :
								break
						for i in range(index+1 , len(content_data)) :
							if(text_count2 == 0) :
								break
							if(mode==2) :
								if(content_data[i].speaker != row.speaker):
									continue
							pinyin = content_data[i].pinyin.split(" ")
							for q in range(0,len(pinyin)) :
								if(text_count2!=0) :
									sub2= sub2 +pinyin[q]+" "
									if(contain_digital(pinyin[q])) :
										text_count2-=1                               
						for i in range(index+1 , len(content_data)) :
							
							if(text_count == 0) :
								break
							if(mode==2) :
								if(content_data[i].speaker != row.speaker):
									continue
							if(find_pos) :							
								post_word = content_data[i].word.split(" ")
								post_word = merge_continuous_english_words(post_word)
								post_pos = content_data[i].pos.split(" ")
								temp_word=""
								my_index = 0
							for q in range(0 ,len(content_data[i].ipu)) :
								if(text_count!=0) :
									if(find_pos) :
										temp_word+= content_data[i].ipu[q] 
										if(post_word[my_index] in temp_word) :
											my_word.append(post_word[my_index])
											my_pos.append(post_pos[my_index])
											my_index+=1
											temp_word=""
									sub += content_data[i].ipu[q] 
									if(is_chinese_string(content_data[i].ipu[q])) :
										text_count-=1

					else :

						text_count = pos
						text_count2 = pos
						sub=""
						sub2=""
						pinyin = row.pinyin
						front = pinyin[0:row.pinyin.find(my_keyword)]
						post = pinyin[row.pinyin.find(my_keyword)+len(my_keyword):len(row.pinyin)]
						post=post.split(" ")
						front= front.split(" ")
						temp = get_my_num(row.pinyin[0:row.pinyin.find(my_keyword)].split(" "))
						chinese_front_num = pinyinnum2chinese(row.ipu,temp)
					
						chinese_front = row.ipu[0:chinese_front_num]
						keyword_count = my_keyword.split(" ")
						chinese_post = row.ipu[chinese_front_num+len(keyword_count) : len(row.ipu)]
						chinese_keyword =row.ipu[chinese_front_num:chinese_front_num+len(keyword_count)]
						if(find_pos):
							front_word = row.word[0:row.word.find(chinese_keyword)].split(" ")							
							front_word = merge_continuous_english_words(front_word)
							post_word = row.word[row.word.find(chinese_keyword)+len(chinese_keyword):len(row.word)].split(" ")
							post_word = merge_continuous_english_words(post_word)	
							front_pos = row.pos.split(" ")
							front_pos2 = row.pos.split(" ")						
							if(front_word):
								if(front_word[len(front_word) -1])=="" :
									front_word.pop()
							my_index = len(front_word)-1
							my_index2 = len(front_word)
							post_pos = row.pos.split(" ")
							post_pos = post_pos[my_index2+1:len(post_pos)]
							my_word = []							
							my_pos=[]
							temp_word=""								
						for i in range(len(chinese_front)-1,-1 ,-1) :

							if(text_count>0) :
								sub = chinese_front[i] + sub
								if(find_pos) :
									temp_word=chinese_front[i] + temp_word
									if(front_word[my_index] in temp_word) :
										my_word.insert(0,front_word[my_index])
										my_pos.insert(0,front_pos[my_index])
										my_index-=1
										temp_word=""
								if(is_chinese_string(chinese_front[i])) :
									text_count-=1                       
						for i in range(index-1 , -1 , -1) :
							if(text_count <= 0) :
								break
							if(mode==2) :
								if(content_data[i].speaker != row.speaker):
									continue
							if(find_pos) :							
								front_word = content_data[i].word.split(" ")
								front_word = merge_continuous_english_words(front_word)
								front_pos = content_data[i].pos.split(" ")
								temp_word=""
								my_index = len(front_word)-1
							for q in range(len(content_data[i].ipu)-1,-1 ,-1) :
								if(text_count!=0) :
									if(find_pos) :
										temp_word= content_data[i].ipu[q] + temp_word
										if(front_word[my_index] in temp_word) :
											my_word.insert(0,front_word[my_index])
											my_pos.insert(0,front_pos[my_index])
											my_index-=1
											temp_word=""
									sub = content_data[i].ipu[q] + sub
									if(is_chinese_string(content_data[i].ipu[q])) :
										text_count-=1
						for i in range(len(front)-1,-1,-1) :
							if(text_count2>0) :
									sub2= front[i] +" "+sub2
									if(contain_digital(front[i])) :
										text_count2-=1
						for i in range(index-1 , -1 , -1) :
							if(text_count2 <= 0) :
								break
							if(mode==2) :
								if(content_data[i].speaker != row.speaker):
									continue
							pinyin2 = content_data[i].pinyin.split(" ")
							for q in range(len(pinyin2)-1,-1 ,-1) :
								if(text_count2!=0) :
									sub2= pinyin2[q] +" "+sub2
									if(contain_digital(pinyin2[q])) :
										text_count2-=1                           
						if(find_pos) :
							my_word.append(chinese_keyword)
							my_pos.append(front_pos2[my_index2])		

						sub2 =sub2 + my_keyword +" "
						for i in range(0,len(keyword_count)) :
							sub+=row.ipu[chinese_front_num+i]
						text_count2 = pos
						text_count = pos
						temp_word = ""
						my_index=0 
						for i in range(0,len(chinese_post)) :
							if(text_count!=0) :
								if(find_pos) :
									temp_word+=chinese_post[i]
									if(post_word[my_index] in temp_word) :
										my_word.append(post_word[my_index])
										my_pos.append(post_pos[my_index])
										my_index+=1
										temp_word=""
								sub += chinese_post[i] 
								if(is_chinese_string(chinese_post[i])) :
									text_count-=1   
						for i in range(index+1 , len(content_data)) :                           
							if(text_count == 0) :
								break
							if(mode==2) :
								if(content_data[i].speaker != row.speaker):
									continue
							if(find_pos) :							
								post_word = content_data[i].word.split(" ")
								post_word = merge_continuous_english_words(post_word)
								post_pos = content_data[i].pos.split(" ")
								temp_word=""
								my_index = 0
							for q in range(0 ,len(content_data[i].ipu)) :
								if(text_count!=0) :
									if(find_pos) :
										temp_word+= content_data[i].ipu[q] 
										if(post_word[my_index] in temp_word) :
											my_word.append(post_word[my_index])
											my_pos.append(post_pos[my_index])
											my_index+=1
											temp_word=""
									sub += content_data[i].ipu[q] 
									if(is_chinese_string(content_data[i].ipu[q])) :
										text_count-=1
						for i in range(0,len(post)) :
							if(text_count2!=0) :
								sub2 = sub2 +post[i] +" "
								if(contain_digital(post[i])) :
									text_count2-=1
							else :
								break
						for i in range(index+1 , len(content_data)) :
							if(text_count2 <= 0) :
								break
							if(mode==2) :
								if(content_data[i].speaker != row.speaker):
									continue
							pinyin = content_data[i].pinyin.split(" ")
							for q in range(0,len(pinyin)) :
								if(text_count2!=0) :
									sub2= sub2 +pinyin[q]+" "
									if(contain_digital(pinyin[q])) :
										text_count2-=1 

						sub2 = remove_extra_spaces(sub2)                                                         
					for each in addtion_search :
							if(each[0] == "and"):
								if(is_chinese_string(each[1])) :
									if(each[1]  in sub) :
										is_matched = True
										if(each[2]!=""):
											try :
												my_index = my_word.index(each[1])
												if(my_pos[my_index] == each[2]) :
													is_matched = True
												else :
													is_matched = False								
											except :
												is_matched = False										
											
									else :
										is_matched = False
										break
								else :
									if(each[1]  in sub2) :
										is_matched = True
										if(each[2]!=""):
											try:
												my_splice = splice(sub,my_word,sub2)
												pinyin_word = pinyin2word(my_word,my_splice)
											except  :
												print(my_word)
												print(my_splice)
												is_matched = False	
											try :
												my_index = pinyin_word.index(each[1])
												if(my_pos[my_index] == each[2]) :
													is_matched = True
												else :

													is_matched = False			
											except :
												print(sub)
												print(my_word)
												print(sub2)
												is_matched = False										
									else :
										is_matched = False
										break

							if(each[0] == "not"):
								if(is_chinese_string(each[1])) :
									if(each[1]  not in sub) :
										is_matched = True
									else :
										is_matched = False
										break
								else :
									if(each[1]  not in sub2) :
										is_matched = True
									else :
										is_matched = False
										break

				if(is_matched) :
					text_count = 20
					start_time = content_data[index].start_time
					temp = []
					temp2 = []
					

					#抓取前20個字
					for i in range(index-1 , 0 , -1) :
						#保持在第一位
						temp.insert(0,content_data[i])
						if(text_count - len(content_data[i].ipu) < 0) :
							break
						text_count = text_count - len(content_data[i].ipu)
					#抓取前一分鐘的字
					for i in range(index-1 , 0 , -1) :
						temp2.insert(0,content_data[i])                  
						if(start_time - content_data[i].start_time > 60) :
							break
					#保留初始數據
					temp.append(content_data[index])
					temp2.append(content_data[index])
					text_count = 20
					#抓取後一分鐘的字
					for i in range(index+1 , len(content_data)) :
						temp2.append(content_data[i])
						if(content_data[i].start_time - start_time >60) :
							break
					#抓取後20個字
					for i in range(index+1 , len(content_data)) :
						temp.append(content_data[i])
						if(text_count - len(content_data[i].ipu) < 0) :
							break
						text_count = text_count - len(content_data[i].ipu)
					my_conversation = Conversation()
					my_conversation.content_array = temp
					my_conversation.sub_content_array = temp2
					my_conversation.talk_id = data_content.talk_id
					my_conversation.theme = data_content.theme
					results.append(my_conversation)
	
	json_data = json.dumps([content.to_json() for content in results], indent=4)
	return JsonResponse({'total': total_count,'serach_num':len(results),'search_corpus':json_data})

@csrf_exempt
def account_detail(request):
	with open('static/media/link.json', 'r', encoding='utf-16') as file:
		json_data = json.load(file)
	db = pymysql.connect(host=json_data['db']['host'],
					user=json_data['db']['user'],
					password=json_data['db']['passwd'],
					database=json_data['db']['db'])
	cursor = db.cursor()
	sql = "SELECT * FROM `Member` JOIN `Member_Log` ON Member.id = Member_Log.Member_id WHERE Member.account='%s';"%request.session['login']
	cursor.execute(sql)
	results = cursor.fetchall()
	data = []
	for row in results:
		data.append({
			'name':row[3],
			'account':row[1],
			'institution':row[6],
			'email':row[4],
			'phone':row[5],
			'start': row[10],
			'end':  row[11]
		})
	return JsonResponse(data, safe=False)

@csrf_exempt
def	Modify_account(request) :
	result = {'status':'','msg':''}
	with open('static/media/link.json', 'r', encoding='utf-16') as file:
		json_data = json.load(file)
	db = pymysql.connect(host=json_data['db']['host'],
                     user=json_data['db']['user'],
                     password=json_data['db']['passwd'],
                     database=json_data['db']['db'])
	cursor = db.cursor()
	try:
		account = request.POST['account']
		email = request.POST['email']
		phone = request.POST['phone']
		institution = request.POST['institution']
		current_datetime = datetime.now()
		date_string = current_datetime.strftime("%Y-%m-%d %H:%M:%S")
		if check_illegal_parameter(email) or check_illegal_parameter(phone) or check_illegal_parameter(institution)  :
			result['status'] = 'No'
			result['msg'] = 'illegal_parameter'		
		else:
			sql = "SELECT * FROM `Member` WHERE `email`='%s'  AND `account`!= '%s';"%(email,account)
			cursor.execute(sql)
			has_email = cursor.fetchall()
			if has_email :
				result['status'] = 'No'
				result['msg'] = "信箱重複"
				return JsonResponse(result, safe=False)
			sql = "SELECT * FROM `Member` WHERE `email`='%s'  AND `account`= '%s';"%(email,account)
			cursor.execute(sql)
			has_email2 = cursor.fetchall()
			if not has_email2 :
				valid  = validate_email(email,check_mx=True ,verify=True,smtp_timeout=5)
				if (valid):
					result['status'] = 'Yes'
					result['msg'] = "信箱可用"
				else :
					result['status'] = 'No'
					result['msg'] = "信箱無效"
					return JsonResponse(result, safe=False)
			sql = "UPDATE `Member` SET  `email`='%s', `phone`='%s', `institution`='%s', `Log_time`='%s' WHERE `account`='%s';"%(email, phone, institution, date_string,account)
			cursor.execute(sql)
			db.commit()			
			result['status'] = 'Yes'
			result['msg'] = '更改成功！'	

	except Exception as e:
		result['status'] = 'Error'
		result['msg'] = str(e)
	db.close()
	record("編輯帳號",request.session['login'],0)
	return JsonResponse(result, safe=False)

@csrf_exempt
def	Modify_password(request) :
	result = {'status':'','msg':''}
	with open('static/media/link.json', 'r', encoding='utf-16') as file:
		json_data = json.load(file)
	db = pymysql.connect(host=json_data['db']['host'],
                     user=json_data['db']['user'],
                     password=json_data['db']['passwd'],
                     database=json_data['db']['db'])
	cursor = db.cursor()
	try:
		account = request.POST['account']
		password = request.POST['password']
		password2 = request.POST['password_new']
		current_datetime = datetime.now()
		date_string = current_datetime.strftime("%Y-%m-%d %H:%M:%S")
		if check_illegal_parameter(password) :
			result['status'] = 'No'
			result['msg'] = 'illegal_parameter'		
		else:
			sql = "SELECT * FROM `Member` WHERE `account`='%s' AND pwd='%s';"%(account,password)
			cursor.execute(sql)
			has_account = cursor.fetchall()
			if(has_account):
				sql = "UPDATE `Member` SET  `pwd`='%s', `Log_time`='%s' WHERE `account`='%s';"%(password, date_string,account)
				print(sql)
				cursor.execute(sql)
				db.commit()			
				result['status'] = 'Yes'
				result['msg'] = '更改成功！'
			else :
				result['status'] = 'No'
				result['msg'] = '密碼錯誤!'

	except Exception as e:
		result['status'] = 'Error'
		result['msg'] = str(e)
	db.close()
	record("變更密碼",request.session['login'],0)
	return JsonResponse(result, safe=False)